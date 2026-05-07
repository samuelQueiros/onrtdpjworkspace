import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from sqlalchemy.orm import Session
from datetime import datetime, date

from app.database import get_db
from app.models.ferias import Ferias
from app.models.log import Log
from app.models.user import User
from app.core.security import require_admin

router = APIRouter(prefix="/importacao", tags=["Importação"])


def _parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.date() if isinstance(value, datetime) else value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


@router.post("/ferias", status_code=status.HTTP_200_OK)
async def importar_ferias(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="Biblioteca openpyxl não instalada no servidor")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Arquivo deve ser .xlsx ou .xls")

    conteudo = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Erro ao ler planilha: {exc}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Planilha vazia")

    # Espera colunas: email_usuario, data_inicio, data_fim (ignorar cabeçalho se for texto)
    header = rows[0]
    data_rows = rows[1:] if isinstance(header[0], str) and not _parse_date(header[0]) else rows

    inseridos = 0
    erros: list[str] = []

    for i, row in enumerate(data_rows, start=2):
        if len(row) < 3:
            erros.append(f"Linha {i}: colunas insuficientes (esperado email, data_inicio, data_fim)")
            continue

        email_val, inicio_val, fim_val = row[0], row[1], row[2]
        ferias_acordo = bool(row[3]) if len(row) > 3 else False

        user = db.query(User).filter(User.email == str(email_val).strip()).first()
        if not user:
            erros.append(f"Linha {i}: usuário '{email_val}' não encontrado")
            continue

        data_inicio = _parse_date(inicio_val)
        data_fim = _parse_date(fim_val)

        if not data_inicio or not data_fim:
            erros.append(f"Linha {i}: datas inválidas ({inicio_val} / {fim_val})")
            continue

        if data_fim < data_inicio:
            erros.append(f"Linha {i}: data_fim anterior a data_inicio")
            continue

        # Evitar duplicatas exatas
        existente = db.query(Ferias).filter(
            Ferias.user_id == user.id,
            Ferias.data_inicio == data_inicio,
            Ferias.data_fim == data_fim,
        ).first()
        if existente:
            erros.append(f"Linha {i}: período {data_inicio}–{data_fim} já existe para {email_val}")
            continue

        dias = (data_fim - data_inicio).days + 1
        nova = Ferias(
            user_id=user.id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            dias_usados=dias,
            status="aprovada",
            ferias_acordo=ferias_acordo if isinstance(ferias_acordo, bool) else False,
        )
        db.add(nova)
        inseridos += 1

    if inseridos:
        log = Log(
            user_id=current_user.id,
            acao="FERIAS_IMPORTADAS",
            detalhes=f"{inseridos} período(s) importado(s) via Excel",
        )
        db.add(log)
        db.commit()

    return {
        "inseridos": inseridos,
        "erros": erros,
        "mensagem": f"{inseridos} registro(s) importado(s) com sucesso. {len(erros)} erro(s).",
    }


@router.post("/logs", status_code=status.HTTP_200_OK)
async def importar_logs(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="Biblioteca openpyxl não instalada no servidor")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Arquivo deve ser .xlsx ou .xls")

    conteudo = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Erro ao ler planilha: {exc}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Planilha vazia")

    header = rows[0]
    data_rows = rows[1:] if isinstance(header[0], str) and not _parse_date(str(header[0])) else rows

    inseridos = 0
    erros: list[str] = []

    for i, row in enumerate(data_rows, start=2):
        if len(row) < 3:
            erros.append(f"Linha {i}: colunas insuficientes (esperado data, acao, detalhes)")
            continue

        data_val, acao_val, detalhes_val = row[0], row[1], row[2]
        email_val = row[3] if len(row) > 3 else None

        criado_em = None
        if data_val:
            if isinstance(data_val, datetime):
                criado_em = data_val
            else:
                try:
                    criado_em = datetime.strptime(str(data_val).strip(), "%d/%m/%Y %H:%M")
                except ValueError:
                    try:
                        criado_em = datetime.strptime(str(data_val).strip(), "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        pass

        user_id = None
        if email_val:
            u = db.query(User).filter(User.email == str(email_val).strip()).first()
            if u:
                user_id = u.id

        log = Log(
            user_id=user_id,
            acao=str(acao_val).strip() if acao_val else "IMPORTADO",
            detalhes=str(detalhes_val).strip() if detalhes_val else None,
            criado_em=criado_em or datetime.utcnow(),
        )
        db.add(log)
        inseridos += 1

    if inseridos:
        db.commit()

    return {
        "inseridos": inseridos,
        "erros": erros,
        "mensagem": f"{inseridos} log(s) importado(s) com sucesso. {len(erros)} erro(s).",
    }
