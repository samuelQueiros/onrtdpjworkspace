from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.repositories import relatorios_repository
from app.services.ferias_service import calcular_saldo, get_ciclo_atual


def formatar_ferias_periodo(ferias) -> dict:
    return {
        "id": ferias.id,
        "data_inicio": ferias.data_inicio,
        "data_fim": ferias.data_fim,
        "dias_usados": ferias.dias_usados,
    }


def formatar_ferias_ciclo(ferias) -> dict:
    response = formatar_ferias_periodo(ferias)
    response.update(
        {
            "status": ferias.status,
            "ferias_acordo": ferias.ferias_acordo,
        }
    )
    return response


def relatorio_colaboradores(db: Session) -> dict:
    colaboradores = []

    for user in relatorios_repository.listar_usuarios_ordenados(db):
        ciclo_inicio, ciclo_fim = get_ciclo_atual(user.data_admissao)
        ferias_ciclo = relatorios_repository.listar_ferias_aprovadas_ciclo(db, user.id, ciclo_inicio)
        ferias_acordo = relatorios_repository.listar_ferias_acordo_aprovadas(db, user.id)
        ferias_pendentes = relatorios_repository.listar_ferias_pendentes_usuario(db, user.id)
        dias_usados = sum(f.dias_usados for f in ferias_ciclo)

        departamento = None
        if user.departamento_id and user.departamento:
            departamento = {"id": user.departamento.id, "nome": user.departamento.nome}

        colaboradores.append(
            {
                "id": user.id,
                "nome": user.nome,
                "email": user.email,
                "departamento": departamento,
                "dias_totais": user.dias_totais,
                "dias_usados": dias_usados,
                "dias_restantes": calcular_saldo(db, user),
                "ciclo_inicio": ciclo_inicio,
                "ciclo_fim": ciclo_fim,
                "ferias": [formatar_ferias_ciclo(f) for f in ferias_ciclo],
                "ferias_acordo": [formatar_ferias_periodo(f) for f in ferias_acordo],
                "ferias_pendentes": [formatar_ferias_periodo(f) for f in ferias_pendentes],
            }
        )

    return {"colaboradores": colaboradores}


def dashboard_admin(db: Session) -> dict:
    hoje = date.today()

    pessoas_em_ferias = []
    for ferias in relatorios_repository.listar_ferias_em_andamento(db, hoje):
        if ferias.usuario:
            pessoas_em_ferias.append(
                {
                    "id": ferias.user_id,
                    "nome": ferias.usuario.nome,
                    "cor": ferias.usuario.cor,
                    "data_inicio": ferias.data_inicio,
                    "data_fim": ferias.data_fim,
                    "dias_restantes": (ferias.data_fim - hoje).days + 1,
                }
            )

    proximas_ferias = []
    for ferias in relatorios_repository.listar_proximas_ferias(db, hoje, hoje + timedelta(days=30)):
        if ferias.usuario:
            proximas_ferias.append(
                {
                    "id": ferias.id,
                    "nome_usuario": ferias.usuario.nome,
                    "cor": ferias.usuario.cor,
                    "data_inicio": ferias.data_inicio,
                    "data_fim": ferias.data_fim,
                    "dias_usados": ferias.dias_usados,
                }
            )

    alertas = []
    for ferias in relatorios_repository.listar_alertas_contabilidade(db, hoje, hoje + timedelta(days=6)):
        if ferias.usuario:
            alertas.append(
                {
                    "ferias_id": ferias.id,
                    "nome_usuario": ferias.usuario.nome,
                    "data_inicio": ferias.data_inicio,
                    "data_fim": ferias.data_fim,
                    "dias_para_inicio": (ferias.data_inicio - hoje).days,
                }
            )

    return {
        "total_colaboradores": relatorios_repository.contar_colaboradores(db),
        "total_ferias_aprovadas": relatorios_repository.contar_ferias_por_status(db, "aprovada"),
        "total_ferias_pendentes": relatorios_repository.contar_ferias_por_status(db, "pendente"),
        "total_autorizacoes_equipamentos_pendentes": (
            relatorios_repository.contar_autorizacoes_equipamentos_por_status(db, "pendente")
        ),
        "total_ferias_rejeitadas": relatorios_repository.contar_ferias_por_status(db, "rejeitada"),
        "total_departamentos": relatorios_repository.contar_departamentos(db),
        "pessoas_em_ferias": pessoas_em_ferias,
        "proximas_ferias": proximas_ferias,
        "alertas_contabilidade": alertas,
    }


def listar_logs(db: Session, page: int = 1, page_size: int = 50) -> dict:
    items = [
        {
            "id": log.id,
            "user_id": log.user_id,
            "nome_usuario": log.usuario.nome if log.usuario else "Sistema",
            "email_usuario": log.usuario.email if log.usuario else None,
            "acao": log.acao,
            "detalhes": log.detalhes,
            "criado_em": log.criado_em,
        }
        for log in relatorios_repository.listar_logs(db, (page - 1) * page_size, page_size)
    ]
    return {"items": items, "page": page, "page_size": page_size, "total": relatorios_repository.contar_logs(db)}


def exportar_logs_xlsx(db: Session) -> bytes:
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Logs"
    planilha.append(["Data", "Ação", "Usuário", "E-mail", "Detalhes"])

    for log in relatorios_repository.listar_todos_logs(db):
        planilha.append([
            log.criado_em,
            log.acao,
            log.usuario.nome if log.usuario else "Sistema",
            log.usuario.email if log.usuario else "",
            log.detalhes or "",
        ])

    preenchimento = PatternFill("solid", fgColor="14213D")
    for celula in planilha[1]:
        celula.fill = preenchimento
        celula.font = Font(color="FFFFFF", bold=True)
        celula.alignment = Alignment(horizontal="center")

    for celula in planilha["A"][1:]:
        if celula.value:
            celula.number_format = "dd/mm/yyyy hh:mm:ss"

    for coluna, largura in zip(("A", "B", "C", "D", "E"), (22, 34, 28, 32, 80)):
        planilha.column_dimensions[coluna].width = largura

    planilha.freeze_panes = "A2"
    planilha.auto_filter.ref = planilha.dimensions

    arquivo = BytesIO()
    workbook.save(arquivo)
    return arquivo.getvalue()
