import io
import re
import unicodedata
import zipfile
from datetime import UTC, date, datetime

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.cpf import validar_cpf
from app.core.crypto import hash_dado_sensivel
from app.models.ferias import Ferias
from app.models.log import Log
from app.models.user import User
from app.repositories import (
    cargos_repository,
    departamentos_repository,
    importacao_repository,
    users_repository,
)
from app.schemas.user import DadosBancarios, Endereco, UserCreate
from app.services import ferias_service, users_service

MAX_IMPORT_SIZE = 5 * 1024 * 1024
MAX_IMPORT_UNCOMPRESSED_SIZE = 50 * 1024 * 1024
MAX_IMPORT_ROWS = 10_000
MAX_IMPORT_COLUMNS = 100

CORES_IMPORTACAO = [
    "#2563EB", "#DC2626", "#16A34A", "#9333EA", "#EA580C", "#0891B2",
    "#DB2777", "#4F46E5", "#65A30D", "#0F766E", "#B45309", "#7C3AED",
]

CABECALHOS_COLABORADORES = [
    "Nome",
    "E-mail",
    "CPF",
    "Cargo",
    "Departamento",
    "Telefone",
    "Contato de emergência 1",
    "Contato de emergência 2",
    "Perfil",
    "Status",
    "Data de admissão",
    "Data de aniversário",
    "Saldo de férias",
    "Dias usados",
    "Próxima concessão",
    "Endereço - Logradouro",
    "Endereço - Número",
    "Endereço - Bairro",
    "Endereço - Cidade",
    "Endereço - CEP",
    "Banco",
    "Agência",
    "Conta",
    "CPF do titular",
    "Nome do titular",
    "Chave PIX",
    "Dias de férias por período",
    "Senha temporária",
]


def _normalizar_cabecalho(value) -> str:
    texto = unicodedata.normalize("NFKD", str(value or ""))
    texto = "".join(char for char in texto if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "_", texto.lower()).strip("_")


def _texto(value) -> str:
    return " ".join(str(value or "").split())


def _inteiro(value, padrao: int | None = None) -> int | None:
    if value is None or str(value).strip() == "":
        return padrao
    if isinstance(value, float) and not value.is_integer():
        raise ValueError(f"valor inteiro inválido: {value}")
    return int(value)


def _perfil(value) -> str:
    normalizado = _normalizar_cabecalho(value)
    if normalizado in {"usuario", "user", "colaborador"}:
        return "user"
    if normalizado in {"administrador", "admin"}:
        return "admin"
    raise ValueError("Perfil deve ser Usuário ou Administrador")


def _status_ativo(value) -> bool:
    normalizado = _normalizar_cabecalho(value or "Ativo")
    if normalizado in {"ativo", "sim", "true", "1"}:
        return True
    if normalizado in {"inativo", "nao", "false", "0"}:
        return False
    raise ValueError("Status deve ser Ativo ou Inativo")


def _cor_automatica(indice: int, cores_em_uso: set[str]) -> str:
    disponiveis = [cor for cor in CORES_IMPORTACAO if cor.lower() not in cores_em_uso]
    if disponiveis:
        cor = disponiveis[indice % len(disponiveis)]
    else:
        cor = CORES_IMPORTACAO[indice % len(CORES_IMPORTACAO)]
    cores_em_uso.add(cor.lower())
    return cor


def gerar_modelo_colaboradores_xlsx(db: Session | None = None) -> bytes:
    workbook = Workbook()
    instrucoes = workbook.active
    instrucoes.title = "Instruções"
    instrucoes.append(["MODELO PARA IMPORTAÇÃO DE COLABORADORES"])
    instrucoes.append(["Preencha uma linha por colaborador na aba Colaboradores."])
    instrucoes.append(["Não altere os nomes dos cabeçalhos. Campos cadastrais e bancários são obrigatórios."])
    instrucoes.append(["Defina uma senha temporária única para cada colaborador. Ela deverá ser trocada no primeiro acesso."])
    instrucoes.append(["A cor de identificação será gerada automaticamente pelo sistema."])
    instrucoes.append(["Cargo e Departamento devem estar previamente cadastrados no sistema."])
    instrucoes.append(["Perfil: Usuário ou Administrador. Status: Ativo ou Inativo."])
    instrucoes.append(["Datas aceitas: DD/MM/AAAA ou AAAA-MM-DD. Dias usados é apenas informativo e não será importado."])
    instrucoes.column_dimensions["A"].width = 110
    instrucoes["A1"].font = Font(bold=True, color="FFFFFF")
    instrucoes["A1"].fill = PatternFill("solid", fgColor="14213D")

    planilha = workbook.create_sheet("Colaboradores")
    planilha.append(CABECALHOS_COLABORADORES)
    planilha.append([
        "Nome de Exemplo", "nome@empresa.com.br", "529.982.247-25", "Analista",
        "Tecnologia", "(61) 99999-9999", "(61) 98888-8888", "(61) 97777-7777",
        "Usuário", "Ativo", date(2024, 1, 10), date(1995, 5, 20), 30, 0,
        date(2027, 1, 10), "Rua Exemplo", "100", "Centro", "Brasília", "70000-000",
        "Banco Exemplo", "0001", "12345-6", "529.982.247-25", "Nome de Exemplo",
        "nome@empresa.com.br", 30, "Defina-uma-senha-unica",
    ])

    preenchimento = PatternFill("solid", fgColor="14213D")
    for celula in planilha[1]:
        celula.fill = preenchimento
        celula.font = Font(color="FFFFFF", bold=True)
        celula.alignment = Alignment(horizontal="center", wrap_text=True)
    for coluna in ("K", "L", "O"):
        planilha[f"{coluna}2"].number_format = "dd/mm/yyyy"
    for indice in range(1, len(CABECALHOS_COLABORADORES) + 1):
        planilha.column_dimensions[planilha.cell(1, indice).column_letter].width = 22
    planilha.column_dimensions["A"].width = 30
    planilha.column_dimensions["B"].width = 32
    planilha.freeze_panes = "A2"
    planilha.auto_filter.ref = f"A1:{planilha.cell(2, len(CABECALHOS_COLABORADORES)).coordinate}"

    perfil = DataValidation(type="list", formula1='"Usuário,Administrador"')
    status = DataValidation(type="list", formula1='"Ativo,Inativo"')
    planilha.add_data_validation(perfil)
    planilha.add_data_validation(status)
    perfil.add("I2:I1000")
    status.add("J2:J1000")

    if db is not None:
        listas = workbook.create_sheet("Cadastros disponíveis")
        listas.append(["Departamentos", "Cargos"])
        departamentos = departamentos_repository.listar_departamentos(db)
        cargos = cargos_repository.listar_cargos(db)
        for indice in range(max(len(departamentos), len(cargos))):
            listas.append([
                departamentos[indice].nome if indice < len(departamentos) else "",
                cargos[indice].nome if indice < len(cargos) else "",
            ])
        listas.column_dimensions["A"].width = 35
        listas.column_dimensions["B"].width = 35

    arquivo = io.BytesIO()
    workbook.save(arquivo)
    return arquivo.getvalue()


def parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.date() if isinstance(value, datetime) else value

    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_datetime(value) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    for fmt in ("%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def validar_extensao_planilha(filename: str | None) -> None:
    if not filename or not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Arquivo deve ser .xlsx")


def parse_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    texto = str(value).strip().lower()
    if texto in {"1", "sim", "s", "true", "verdadeiro"}:
        return True
    if texto in {"0", "nao", "não", "n", "false", "falso", ""}:
        return False
    raise ValueError(f"valor booleano invalido: {value}")


def carregar_linhas_planilha(conteudo: bytes) -> list[tuple]:
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Biblioteca openpyxl nao instalada no servidor") from exc

    wb = None
    try:
        with zipfile.ZipFile(io.BytesIO(conteudo)) as arquivo_zip:
            tamanho_descompactado = sum(item.file_size for item in arquivo_zip.infolist())
            if tamanho_descompactado > MAX_IMPORT_UNCOMPRESSED_SIZE:
                raise HTTPException(
                    status_code=400,
                    detail="Planilha excede o limite descompactado de 50 MB",
                )
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
        ws = wb["Colaboradores"] if "Colaboradores" in wb.sheetnames else wb.active
        if ws.max_column > MAX_IMPORT_COLUMNS:
            raise HTTPException(status_code=400, detail="Planilha possui colunas demais")
        rows = []
        for indice, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if indice > MAX_IMPORT_ROWS:
                raise HTTPException(
                    status_code=400,
                    detail=f"Planilha excede o limite de {MAX_IMPORT_ROWS} linhas",
                )
            rows.append(row)
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=400, detail="Planilha XLSX invalida ou corrompida") from None
    finally:
        if wb is not None:
            wb.close()
    if not rows:
        raise HTTPException(status_code=400, detail="Planilha vazia")
    return rows


def ignorar_cabecalho(rows: list[tuple]) -> list[tuple]:
    header = rows[0]
    return rows[1:] if header and isinstance(header[0], str) and not parse_date(header[0]) else rows


def _erro_validacao(exc: ValidationError) -> str:
    mensagens = []
    for erro in exc.errors():
        campo = ".".join(str(parte) for parte in erro["loc"])
        mensagens.append(f"{campo}: {erro['msg']}")
    return "; ".join(mensagens)


def importar_colaboradores(
    db: Session,
    filename: str | None,
    conteudo: bytes,
    current_user: User,
) -> dict:
    validar_extensao_planilha(filename)
    rows = carregar_linhas_planilha(conteudo)
    if len(rows) < 2:
        return {
            "inseridos": 0,
            "erros": ["A planilha não possui colaboradores para importar."],
            "mensagem": "Nenhum colaborador importado. Corrija a planilha e tente novamente.",
        }

    cabecalhos = {
        _normalizar_cabecalho(valor): indice
        for indice, valor in enumerate(rows[0])
        if valor is not None
    }
    obrigatorios = {
        "nome", "e_mail", "cpf", "cargo", "departamento", "telefone",
        "contato_de_emergencia_1", "contato_de_emergencia_2", "perfil",
        "data_de_admissao", "data_de_aniversario", "saldo_de_ferias",
        "endereco_logradouro", "endereco_numero", "endereco_bairro",
        "endereco_cidade", "endereco_cep", "banco", "agencia", "conta",
        "cpf_do_titular", "nome_do_titular", "chave_pix",
        "senha_temporaria",
    }
    faltantes = sorted(obrigatorios - set(cabecalhos))
    if faltantes:
        nomes = ", ".join(nome.replace("_", " ") for nome in faltantes)
        raise HTTPException(status_code=400, detail=f"Colunas obrigatórias ausentes: {nomes}")

    def valor(row: tuple, nome: str, padrao=None):
        indice = cabecalhos.get(nome)
        return row[indice] if indice is not None and indice < len(row) else padrao

    cores_em_uso = {
        usuario.cor.lower()
        for usuario in users_repository.listar_usuarios(db)
        if usuario.cor
    }
    emails_planilha: set[str] = set()
    cpfs_planilha: set[str] = set()
    senhas_planilha: set[str] = set()
    preparados: list[tuple[int, UserCreate, bool]] = []
    erros: list[str] = []

    for numero_linha, row in enumerate(rows[1:], start=2):
        if not any(celula is not None and str(celula).strip() for celula in row):
            continue
        try:
            email = _texto(valor(row, "e_mail")).lower()
            cpf_normalizado = validar_cpf(_texto(valor(row, "cpf")))
            senha_temporaria = str(valor(row, "senha_temporaria") or "").strip()
            if email in emails_planilha:
                raise ValueError(f"E-mail repetido na planilha: {email}")
            if cpf_normalizado in cpfs_planilha:
                raise ValueError("CPF repetido na planilha")
            if senha_temporaria in senhas_planilha:
                raise ValueError("Senha temporária repetida na planilha")
            if users_repository.obter_usuario_por_email(db, email):
                raise ValueError(f"E-mail já cadastrado: {email}")
            if users_repository.obter_usuario_por_cpf_hash(db, hash_dado_sensivel(cpf_normalizado)):
                raise ValueError("CPF já cadastrado para outro colaborador")

            departamento_nome = _texto(valor(row, "departamento"))
            departamento = departamentos_repository.obter_departamento_por_nome(db, departamento_nome)
            if not departamento:
                raise ValueError(f"Departamento não cadastrado: {departamento_nome}")
            cargo_nome = _texto(valor(row, "cargo"))
            if not cargos_repository.obter_cargo_por_nome(db, cargo_nome):
                raise ValueError(f"Cargo não cadastrado: {cargo_nome}")

            data_admissao = parse_date(valor(row, "data_de_admissao"))
            data_aniversario = parse_date(valor(row, "data_de_aniversario"))
            proxima_concessao_valor = valor(row, "proxima_concessao")
            proxima_concessao = parse_date(proxima_concessao_valor)
            if not data_admissao:
                raise ValueError("Data de admissão inválida")
            if not data_aniversario:
                raise ValueError("Data de aniversário inválida")
            if proxima_concessao_valor and not proxima_concessao:
                raise ValueError("Próxima concessão inválida")

            payload = UserCreate(
                nome=_texto(valor(row, "nome")),
                email=email,
                cpf=cpf_normalizado,
                senha=senha_temporaria,
                role=_perfil(valor(row, "perfil")),
                dias_totais=_inteiro(valor(row, "dias_de_ferias_por_periodo"), 30),
                saldo_inicial_dias=_inteiro(valor(row, "saldo_de_ferias"), 0),
                proxima_concessao_ferias=proxima_concessao,
                departamento_id=departamento.id,
                data_admissao=data_admissao,
                data_aniversario=data_aniversario,
                cor=_cor_automatica(len(preparados), cores_em_uso),
                telefone=_texto(valor(row, "telefone")),
                telefone_emergencia=_texto(valor(row, "contato_de_emergencia_1")),
                telefone_emergencia_2=_texto(valor(row, "contato_de_emergencia_2")),
                endereco=Endereco(
                    logradouro=valor(row, "endereco_logradouro"),
                    numero=valor(row, "endereco_numero"),
                    bairro=valor(row, "endereco_bairro"),
                    cidade=valor(row, "endereco_cidade"),
                    cep=valor(row, "endereco_cep"),
                ),
                dados_bancarios=DadosBancarios(
                    banco=valor(row, "banco"),
                    agencia=valor(row, "agencia"),
                    conta=valor(row, "conta"),
                    cpf_titular=valor(row, "cpf_do_titular"),
                    nome_titular=valor(row, "nome_do_titular"),
                    chave_pix=valor(row, "chave_pix"),
                ),
                cargo=cargo_nome,
            )
            ativo = _status_ativo(valor(row, "status", "Ativo"))
            emails_planilha.add(email)
            cpfs_planilha.add(cpf_normalizado)
            senhas_planilha.add(senha_temporaria)
            preparados.append((numero_linha, payload, ativo))
        except ValidationError as exc:
            erros.append(f"Linha {numero_linha}: {_erro_validacao(exc)}")
        except (TypeError, ValueError) as exc:
            erros.append(f"Linha {numero_linha}: {exc}")

    if erros:
        return {
            "inseridos": 0,
            "erros": erros,
            "mensagem": (
                f"Nenhum colaborador foi importado. "
                f"Corrija {len(erros)} erro(s) e envie a planilha novamente."
            ),
        }
    if not preparados:
        return {
            "inseridos": 0,
            "erros": ["A planilha não possui colaboradores para importar."],
            "mensagem": "Nenhum colaborador importado.",
        }

    inseridos = 0
    try:
        for _, payload, ativo in preparados:
            users_service.criar_usuario(db, payload, current_user, commit=False)
            if not ativo:
                usuario = users_repository.obter_usuario_por_email(db, payload.email)
                usuario.ativo = False
            inseridos += 1
        db.add(Log(
            user_id=current_user.id,
            acao="COLABORADORES_IMPORTADOS",
            detalhes=(
                f"{inseridos} colaborador(es) importado(s) via planilha. "
                "Cores geradas automaticamente e troca de senha obrigatória ativada."
            ),
        ))
        db.commit()
    except (HTTPException, IntegrityError) as exc:
        db.rollback()
        detalhe = exc.detail if isinstance(exc, HTTPException) else "E-mail ou CPF duplicado"
        raise HTTPException(
            status_code=409,
            detail=f"Falha durante a importação após {inseridos} registro(s): {detalhe}",
        ) from exc

    return {
        "inseridos": inseridos,
        "erros": [],
        "mensagem": (
            f"{inseridos} colaborador(es) importado(s) com sucesso. "
            "As senhas temporárias devem ser trocadas no primeiro acesso."
        ),
    }


def importar_ferias(db: Session, filename: str | None, conteudo: bytes, current_user: User) -> dict:
    validar_extensao_planilha(filename)
    data_rows = ignorar_cabecalho(carregar_linhas_planilha(conteudo))

    inseridos = 0
    erros: list[str] = []

    for i, row in enumerate(data_rows, start=2):
        if len(row) < 3:
            erros.append(f"Linha {i}: colunas insuficientes (esperado email, data_inicio, data_fim)")
            continue

        email_val, inicio_val, fim_val = row[0], row[1], row[2]
        try:
            ferias_acordo = parse_bool(row[3] if len(row) > 3 else None)
        except ValueError as exc:
            erros.append(f"Linha {i}: {exc}")
            continue

        user = importacao_repository.obter_usuario_por_email(db, str(email_val).strip())
        if not user:
            erros.append(f"Linha {i}: usuario '{email_val}' nao encontrado")
            continue

        data_inicio = parse_date(inicio_val)
        data_fim = parse_date(fim_val)
        if not data_inicio or not data_fim:
            erros.append(f"Linha {i}: datas invalidas ({inicio_val} / {fim_val})")
            continue

        if data_fim < data_inicio:
            erros.append(f"Linha {i}: data_fim anterior a data_inicio")
            continue

        if importacao_repository.existe_ferias_periodo(db, user.id, data_inicio, data_fim):
            erros.append(f"Linha {i}: periodo {data_inicio}-{data_fim} ja existe para {email_val}")
            continue

        try:
            ferias_service.bloquear_regras_ferias(db, user)
            ferias_service.garantir_saldo_atualizado(
                db,
                user,
                current_user.id,
                commit=False,
            )
            ferias_service.verificar_regras_data(data_inicio, data_fim)
            ferias_service.verificar_bloqueio_datas(db, data_inicio, data_fim)
            ferias_service.verificar_sobreposicao_usuario(
                db,
                user.id,
                data_inicio,
                data_fim,
            )
            dias = ferias_service.calcular_dias(data_inicio, data_fim)
            if not ferias_acordo:
                if ferias_service.verificar_sobreposicao_departamento(db, user, data_inicio, data_fim):
                    raise HTTPException(status_code=400, detail="limite simultaneo do departamento atingido")
                saldo = ferias_service.calcular_saldo(db, user)
                if saldo < dias:
                    raise HTTPException(status_code=400, detail=f"saldo insuficiente ({saldo} dias)")
        except HTTPException as exc:
            db.rollback()
            erros.append(f"Linha {i}: {exc.detail}")
            continue
        importacao_repository.adicionar_ferias(
            db,
            Ferias(
                user_id=user.id,
                data_inicio=data_inicio,
                data_fim=data_fim,
                dias_usados=dias,
                status="aprovada",
                ferias_acordo=ferias_acordo,
            ),
        )
        importacao_repository.adicionar_log(
            db,
            Log(
                user_id=current_user.id,
                acao="FERIAS_IMPORTADA",
                detalhes=f"Periodo de {data_inicio} a {data_fim} importado para usuario #{user.id}",
            ),
        )
        # Cada linha usa sua propria transacao para liberar advisory locks e
        # impedir deadlocks entre lotes processados em ordens diferentes.
        try:
            importacao_repository.commit(db)
            inseridos += 1
        except IntegrityError:
            db.rollback()
            erros.append(f"Linha {i}: conflito com outro registro gravado simultaneamente")

    return {
        "inseridos": inseridos,
        "erros": erros,
        "mensagem": f"{inseridos} registro(s) importado(s) com sucesso. {len(erros)} erro(s).",
    }


def importar_logs(db: Session, filename: str | None, conteudo: bytes, current_user: User) -> dict:
    validar_extensao_planilha(filename)
    data_rows = ignorar_cabecalho(carregar_linhas_planilha(conteudo))

    inseridos = 0
    erros: list[str] = []

    for i, row in enumerate(data_rows, start=2):
        if len(row) < 3:
            erros.append(f"Linha {i}: colunas insuficientes (esperado data, acao, detalhes)")
            continue

        data_val, acao_val, detalhes_val = row[0], row[1], row[2]
        email_val = row[3] if len(row) > 3 else None

        user_id = None
        if email_val:
            user = importacao_repository.obter_usuario_por_email(db, str(email_val).strip())
            if not user:
                erros.append(f"Linha {i}: usuario '{email_val}' nao encontrado")
                continue
            user_id = user.id

        criado_em = parse_datetime(data_val)
        if criado_em is None:
            erros.append(f"Linha {i}: data invalida ({data_val})")
            continue

        acao_original = str(acao_val).strip() if acao_val else "SEM_ACAO"
        detalhes_originais = str(detalhes_val).strip() if detalhes_val else ""

        importacao_repository.adicionar_log(
            db,
            Log(
                user_id=user_id,
                acao=f"IMPORTADO::{acao_original}",
                detalhes=(
                    f"[Importado pelo administrador #{current_user.id}] "
                    f"{detalhes_originais}"
                ).strip(),
                criado_em=criado_em,
            ),
        )
        inseridos += 1

    if inseridos:
        importacao_repository.adicionar_log(
            db,
            Log(
                user_id=current_user.id,
                acao="LOTE_LOGS_IMPORTADO",
                detalhes=(
                    f"Lote com {inseridos} log(s) historico(s) importado. "
                    f"{len(erros)} linha(s) rejeitada(s)."
                ),
                criado_em=datetime.now(UTC).replace(tzinfo=None),
            ),
        )
        importacao_repository.commit(db)

    return {
        "inseridos": inseridos,
        "erros": erros,
        "mensagem": f"{inseridos} log(s) importado(s) com sucesso. {len(erros)} erro(s).",
    }
