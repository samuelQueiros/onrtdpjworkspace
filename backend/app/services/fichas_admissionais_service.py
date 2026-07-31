from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError
from sqlalchemy.orm import Session

from app.core.crypto import criptografar_dado_sensivel, descriptografar_dado_sensivel
from app.core.timezone import hoje_sao_paulo
from app.models.ficha_admissional import FichaAdmissional
from app.models.historico_salarial import HistoricoSalarial
from app.models.log import Log
from app.models.user import User
from app.repositories import fichas_admissionais_repository, historico_salarial_repository
from app.schemas.ficha_admissional import FichaAdmissionalUpdate
from app.services import historico_colaborador_service, importacao_service, users_service


CAMPOS_CRIPTOGRAFADOS = {
    "local_nascimento": "local_nascimento_criptografado",
    "nacionalidade": "nacionalidade_criptografada",
    "nome_mae": "nome_mae_criptografado",
    "nome_pai": "nome_pai_criptografado",
    "pis_numero": "pis_numero_criptografado",
    "rg_numero": "rg_numero_criptografado",
    "rg_orgao_emissor": "rg_orgao_emissor_criptografado",
    "ctps_numero": "ctps_numero_criptografado",
    "ctps_serie": "ctps_serie_criptografada",
    "telefone_alternativo": "telefone_alternativo_criptografado",
    "email_alternativo": "email_alternativo_criptografado",
    "nome_conjuge": "nome_conjuge_criptografado",
    "salario": "salario_criptografado",
    "horario_trabalho": "horario_trabalho_criptografado",
    "dias_semana": "dias_semana_criptografados",
    "vale_transporte": "vale_transporte_criptografado",
    "beneficios": "beneficios_criptografados",
    "valor_beneficios": "valor_beneficios_criptografado",
}

CAMPOS_DIRETOS = (
    "uf_nascimento",
    "sexo",
    "pis_emissao",
    "rg_emissao",
    "ctps_validade",
    "ctps_uf",
    "ctps_emissao",
    "endereco_uf",
    "estado_civil",
    "grau_instrucao",
    "contrato_experiencia_dias",
    "status",
)

CAMPOS_MODELO = (
    ("Informações do trabalhador", "Local do nascimento", "local_nascimento", "Cidade de nascimento", "texto"),
    ("Informações do trabalhador", "UF de nascimento", "uf_nascimento", "Sigla com duas letras", "uf"),
    ("Informações do trabalhador", "Nacionalidade", "nacionalidade", "Ex.: Brasileira", "texto"),
    ("Informações do trabalhador", "Sexo", "sexo", "Feminino, Masculino, Outro ou Não informado", "sexo"),
    ("Informações do trabalhador", "Nome da mãe", "nome_mae", "Nome completo, sem abreviações", "texto"),
    ("Informações do trabalhador", "Nome do pai", "nome_pai", "Nome completo, sem abreviações", "texto"),
    ("Documentos", "Número do PIS", "pis_numero", "Manter zeros à esquerda", "texto"),
    ("Documentos", "Emissão do PIS", "pis_emissao", "DD/MM/AAAA", "data"),
    ("Documentos", "Número do RG", "rg_numero", "Incluir dígito, quando houver", "texto"),
    ("Documentos", "Emissão do RG", "rg_emissao", "DD/MM/AAAA", "data"),
    ("Documentos", "Órgão emissor do RG", "rg_orgao_emissor", "Ex.: SSP/DF", "texto"),
    ("Documentos", "Número da CTPS", "ctps_numero", "Carteira de Trabalho", "texto"),
    ("Documentos", "Série da CTPS", "ctps_serie", "Manter zeros à esquerda", "texto"),
    ("Documentos", "Validade da CTPS", "ctps_validade", "Quando houver", "data"),
    ("Documentos", "UF da CTPS", "ctps_uf", "Sigla com duas letras", "uf"),
    ("Documentos", "Emissão da CTPS", "ctps_emissao", "DD/MM/AAAA", "data"),
    ("Contato e dados sociais", "Telefone alternativo", "telefone_alternativo", "Com DDD", "texto"),
    ("Contato e dados sociais", "E-mail alternativo", "email_alternativo", "E-mail pessoal", "texto"),
    ("Contato e dados sociais", "UF do endereço", "endereco_uf", "Sigla com duas letras", "uf"),
    ("Contato e dados sociais", "Estado civil", "estado_civil", "Selecione uma opção", "estado_civil"),
    ("Contato e dados sociais", "Nome do cônjuge", "nome_conjuge", "Quando houver", "texto"),
    ("Contato e dados sociais", "Grau de instrução", "grau_instrucao", "Ex.: Ensino Superior Completo", "texto"),
    ("Dados funcionais", "Salário", "salario", "Valor em reais", "moeda"),
    ("Dados funcionais", "Horário de trabalho", "horario_trabalho", "Ex.: 08:00 às 17:00", "texto"),
    ("Dados funcionais", "Dias da semana", "dias_semana", "Ex.: Segunda a sexta-feira", "texto"),
    ("Dados funcionais", "Valor do vale-transporte", "vale_transporte", "Informe zero quando não houver", "moeda"),
    ("Dados funcionais", "Benefícios", "beneficios", "Ex.: Caju, plano de saúde", "texto"),
    ("Dados funcionais", "Valor dos benefícios", "valor_beneficios", "Valor mensal em reais", "moeda"),
    ("Dados funcionais", "Contrato de experiência (dias)", "contrato_experiencia_dias", "Prazo total em dias", "inteiro"),
    ("Controle", "Status da ficha", "status", "Rascunho ou Completa", "status"),
)


ROTULOS_CAMPO_FICHA = {campo: rotulo for _, rotulo, campo, _, _ in CAMPOS_MODELO}


def _valor_criptografado(valor) -> str | None:
    if valor is None or valor == "":
        return None
    return criptografar_dado_sensivel(str(valor))


def _valor_decimal(valor: str | None) -> Decimal | None:
    return Decimal(valor) if valor not in {None, ""} else None


def formatar_ficha(ficha: FichaAdmissional) -> dict:
    resultado = {
        "id": ficha.id,
        "user_id": ficha.user_id,
        "criado_por_id": ficha.criado_por_id,
        "atualizado_por_id": ficha.atualizado_por_id,
        "criado_em": ficha.criado_em,
        "atualizado_em": ficha.atualizado_em,
    }
    for campo, atributo in CAMPOS_CRIPTOGRAFADOS.items():
        valor = descriptografar_dado_sensivel(getattr(ficha, atributo))
        resultado[campo] = _valor_decimal(valor) if campo in {
            "salario", "vale_transporte", "valor_beneficios"
        } else valor
    for campo in CAMPOS_DIRETOS:
        resultado[campo] = getattr(ficha, campo)
    return resultado


def consultar_ficha(db: Session, user_id: int, current_user: User) -> dict | None:
    user = users_service.buscar_usuario(db, user_id)
    ficha = fichas_admissionais_repository.obter_por_usuario(db, user_id)
    db.add(Log(
        user_id=current_user.id,
        acao="FICHA_ADMISSIONAL_CONSULTADA",
        detalhes=f"Ficha admissional de {user.nome} consultada por administrador",
    ))
    db.commit()
    return formatar_ficha(ficha) if ficha else None


def minha_ficha(db: Session, current_user: User) -> dict | None:
    ficha = fichas_admissionais_repository.obter_por_usuario(db, current_user.id)
    return formatar_ficha(ficha) if ficha else None


def atualizar_ficha(
    db: Session,
    user_id: int,
    payload: FichaAdmissionalUpdate,
    current_user: User,
    acao: str = "FICHA_ADMISSIONAL_ATUALIZADA",
    origem_importacao: bool = False,
) -> dict:
    user = users_service.buscar_usuario(db, user_id)
    ficha = fichas_admissionais_repository.obter_por_usuario(db, user_id)
    salario_anterior = (
        _valor_decimal(descriptografar_dado_sensivel(ficha.salario_criptografado)) if ficha else None
    )
    valor_beneficios_anterior = (
        _valor_decimal(descriptografar_dado_sensivel(ficha.valor_beneficios_criptografado)) if ficha else None
    )
    # A criptografia não é determinística (mesmo texto gera cifra diferente
    # a cada vez), então precisa descriptografar antes/depois pra comparar o
    # conteúdo de verdade — o formulário reenvia todos os campos da ficha em
    # todo submit, mesmo os que o admin não tocou.
    valores_anteriores_criptografados = {
        campo: (descriptografar_dado_sensivel(getattr(ficha, atributo)) if ficha else None)
        for campo, atributo in CAMPOS_CRIPTOGRAFADOS.items()
    }
    valores_anteriores_diretos = {
        campo: (getattr(ficha, campo) if ficha else None)
        for campo in CAMPOS_DIRETOS
    }
    if ficha is None:
        ficha = FichaAdmissional(
            user_id=user_id,
            criado_por_id=current_user.id,
            atualizado_por_id=current_user.id,
        )

    for campo, atributo in CAMPOS_CRIPTOGRAFADOS.items():
        if campo in payload.model_fields_set:
            setattr(ficha, atributo, _valor_criptografado(getattr(payload, campo)))
    for campo in CAMPOS_DIRETOS:
        if campo in payload.model_fields_set:
            setattr(ficha, campo, getattr(payload, campo))
    ficha.atualizado_por_id = current_user.id

    fichas_admissionais_repository.salvar(db, ficha, commit=False)

    alteracoes = []
    for campo, atributo in CAMPOS_CRIPTOGRAFADOS.items():
        anterior = valores_anteriores_criptografados[campo]
        novo = descriptografar_dado_sensivel(getattr(ficha, atributo))
        if novo == anterior:
            continue
        rotulo = ROTULOS_CAMPO_FICHA.get(campo, campo)
        alteracoes.append(f"{rotulo}: {anterior or '—'} -> {novo or '—'}")
    for campo in CAMPOS_DIRETOS:
        anterior = valores_anteriores_diretos[campo]
        novo = getattr(ficha, campo)
        if novo != anterior:
            rotulo = ROTULOS_CAMPO_FICHA.get(campo, campo)
            alteracoes.append(f"{rotulo}: {anterior if anterior is not None else '—'} -> {novo if novo is not None else '—'}")
    resumo = "; ".join(alteracoes) if alteracoes else "nenhum campo alterado"

    db.add(Log(
        user_id=current_user.id,
        acao=acao,
        detalhes=f"Ficha admissional de {user.nome} atualizada por administrador: {resumo}",
    ))

    if "salario" in payload.model_fields_set and payload.salario is not None and payload.salario != salario_anterior:
        eh_edicao_de_salario_existente = salario_anterior is not None and not origem_importacao
        if eh_edicao_de_salario_existente and not (payload.motivo_alteracao_salario and payload.tipo_alteracao_salario):
            raise HTTPException(
                status_code=400,
                detail="Informe o motivo e o tipo (reajuste ou correção) da alteração salarial.",
            )
        if origem_importacao:
            tipo, motivo = "reajuste", "Importação de planilha"
        elif salario_anterior is None:
            tipo, motivo = "reajuste", "Cadastro inicial"
        else:
            tipo, motivo = payload.tipo_alteracao_salario, payload.motivo_alteracao_salario
        db.add(HistoricoSalarial(
            user_id=user_id,
            salario_criptografado=criptografar_dado_sensivel(str(payload.salario)),
            data_vigencia=hoje_sao_paulo(),
            tipo=tipo,
            motivo=motivo,
            criado_por_id=current_user.id,
        ))

    if "valor_beneficios" in payload.model_fields_set:
        historico_colaborador_service.processar_alteracao_se_necessario(
            db,
            user_id,
            "valor_beneficios",
            str(valor_beneficios_anterior) if valor_beneficios_anterior is not None else None,
            str(payload.valor_beneficios) if payload.valor_beneficios is not None else None,
            payload.motivo_alteracao_beneficios,
            payload.tipo_alteracao_beneficios,
            current_user,
            eh_importacao=origem_importacao,
        )

    db.commit()
    db.refresh(ficha)
    return formatar_ficha(ficha)


def gerar_modelo_xlsx(db: Session, user_id: int, current_user: User) -> bytes:
    user = users_service.buscar_usuario(db, user_id)
    ficha = fichas_admissionais_repository.obter_por_usuario(db, user_id)
    dados = formatar_ficha(ficha) if ficha else {"status": "rascunho"}

    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Ficha admissional"
    planilha.sheet_view.showGridLines = False
    planilha.merge_cells("A1:C1")
    planilha["A1"] = "FICHA ADMISSIONAL — DADOS COMPLEMENTARES"
    planilha["A1"].font = Font(size=15, bold=True, color="FFFFFF")
    planilha["A1"].fill = PatternFill("solid", fgColor="14213D")
    planilha["A1"].alignment = Alignment(horizontal="center")
    planilha.row_dimensions[1].height = 28
    planilha.append(["Colaborador", user.nome, "Campo de identificação; não será alterado na importação."])
    planilha.append(["E-mail do colaborador", user.email, "O arquivo só poderá ser importado para este e-mail."])
    planilha.append(["Orientação", "Preencha a coluna Valor", "Não altere os nomes dos campos. Campos vazios serão removidos da ficha."])
    planilha.append([])

    linha_por_campo = {}
    linhas_secao = set()
    secao_atual = None
    for secao, rotulo, campo, orientacao, tipo in CAMPOS_MODELO:
        if secao != secao_atual:
            planilha.append([secao.upper(), None, None])
            linha_secao = planilha.max_row
            linhas_secao.add(linha_secao)
            planilha.merge_cells(start_row=linha_secao, start_column=1, end_row=linha_secao, end_column=3)
            celula = planilha.cell(linha_secao, 1)
            celula.fill = PatternFill("solid", fgColor="DCE6F1")
            celula.font = Font(bold=True, color="14213D")
            secao_atual = secao
        valor = dados.get(campo)
        if isinstance(valor, Decimal):
            valor = float(valor)
        planilha.append([rotulo, valor, orientacao])
        linha = planilha.max_row
        linha_por_campo[campo] = linha
        if tipo == "data":
            planilha.cell(linha, 2).number_format = "dd/mm/yyyy"
        elif tipo == "moeda":
            planilha.cell(linha, 2).number_format = '"R$" #,##0.00'
        else:
            planilha.cell(linha, 2).number_format = "@"

    planilha.column_dimensions["A"].width = 34
    planilha.column_dimensions["B"].width = 36
    planilha.column_dimensions["C"].width = 62
    planilha.freeze_panes = "A6"
    for linha in range(2, planilha.max_row + 1):
        if linha in linhas_secao:
            continue
        planilha.cell(linha, 1).font = Font(bold=True)
        planilha.cell(linha, 2).fill = PatternFill("solid", fgColor="FFF7D6")
        planilha.cell(linha, 3).font = Font(color="64748B", italic=True)
        for coluna in range(1, 4):
            planilha.cell(linha, coluna).alignment = Alignment(vertical="top", wrap_text=True)

    validacoes = {
        "sexo": '"Feminino,Masculino,Outro,Não informado"',
        "estado_civil": '"Solteiro,Casado,Divorciado,Viúvo,Separado,União estável,Outro"',
        "status": '"Rascunho,Completa"',
    }
    for campo, formula in validacoes.items():
        validacao = DataValidation(type="list", formula1=formula)
        planilha.add_data_validation(validacao)
        validacao.add(planilha.cell(linha_por_campo[campo], 2))

    db.add(Log(
        user_id=current_user.id,
        acao="MODELO_FICHA_ADMISSIONAL_EXPORTADO",
        detalhes=f"Modelo preenchivel da ficha admissional de {user.nome} exportado",
    ))
    db.commit()
    arquivo = BytesIO()
    workbook.save(arquivo)
    return arquivo.getvalue()


def _texto(valor) -> str | None:
    if valor is None:
        return None
    texto = " ".join(str(valor).split())
    return texto or None


def _data(valor, rotulo: str) -> date | None:
    if valor is None or str(valor).strip() == "":
        return None
    resultado = importacao_service.parse_date(valor)
    if not resultado:
        raise ValueError(f"{rotulo}: data inválida")
    return resultado


def _decimal(valor, rotulo: str) -> Decimal | None:
    if valor is None or str(valor).strip() == "":
        return None
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    texto = str(valor).strip().replace("R$", "").replace(" ", "")
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return Decimal(texto)
    except InvalidOperation as exc:
        raise ValueError(f"{rotulo}: valor monetário inválido") from exc


def _inteiro(valor, rotulo: str) -> int | None:
    if valor is None or str(valor).strip() == "":
        return None
    if isinstance(valor, float) and not valor.is_integer():
        raise ValueError(f"{rotulo}: deve ser um número inteiro")
    try:
        return int(valor)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{rotulo}: deve ser um número inteiro") from exc


def _opcao(valor, opcoes: dict[str, str], rotulo: str) -> str | None:
    if valor is None or str(valor).strip() == "":
        return None
    normalizado = importacao_service._normalizar_cabecalho(valor)
    if normalizado not in opcoes:
        raise ValueError(f"{rotulo}: opção inválida")
    return opcoes[normalizado]


def importar_xlsx(
    db: Session,
    user_id: int,
    filename: str | None,
    conteudo: bytes,
    current_user: User,
) -> dict:
    importacao_service.validar_extensao_planilha(filename)
    user = users_service.buscar_usuario(db, user_id)
    linhas = importacao_service.carregar_linhas_planilha(conteudo, "Ficha admissional")
    valores = {
        importacao_service._normalizar_cabecalho(linha[0]): linha[1] if len(linha) > 1 else None
        for linha in linhas
        if linha and linha[0]
    }

    email_planilha = _texto(valores.get("e_mail_do_colaborador"))
    if not email_planilha or email_planilha.lower() != user.email.lower():
        raise HTTPException(
            status_code=400,
            detail="O e-mail da planilha não corresponde ao colaborador selecionado",
        )

    esperados = {
        importacao_service._normalizar_cabecalho(rotulo)
        for _, rotulo, _, _, _ in CAMPOS_MODELO
    }
    faltantes = esperados - set(valores)
    if faltantes:
        raise HTTPException(
            status_code=400,
            detail="A planilha não corresponde ao modelo de ficha admissional atual",
        )

    sexo = {"feminino": "feminino", "masculino": "masculino", "outro": "outro", "nao_informado": "nao_informado"}
    estado_civil = {
        "solteiro": "solteiro", "casado": "casado", "divorciado": "divorciado",
        "viuvo": "viuvo", "separado": "separado", "uniao_estavel": "uniao_estavel", "outro": "outro",
    }
    status = {"rascunho": "rascunho", "completa": "completa"}
    try:
        dados = {}
        for _, rotulo, campo, _, tipo in CAMPOS_MODELO:
            valor = valores[importacao_service._normalizar_cabecalho(rotulo)]
            if tipo == "data":
                dados[campo] = _data(valor, rotulo)
            elif tipo == "moeda":
                dados[campo] = _decimal(valor, rotulo)
            elif tipo == "inteiro":
                dados[campo] = _inteiro(valor, rotulo)
            elif tipo == "sexo":
                dados[campo] = _opcao(valor, sexo, rotulo)
            elif tipo == "estado_civil":
                dados[campo] = _opcao(valor, estado_civil, rotulo)
            elif tipo == "status":
                dados[campo] = _opcao(valor or "Rascunho", status, rotulo)
            else:
                dados[campo] = _texto(valor)
        payload = FichaAdmissionalUpdate(**dados)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except ValidationError as exc:
        primeiro = exc.errors()[0]
        campo = ".".join(str(item) for item in primeiro["loc"])
        raise HTTPException(
            status_code=400,
            detail=f"Campo {campo}: {primeiro['msg']}",
        ) from exc
    ficha = atualizar_ficha(
        db,
        user_id,
        payload,
        current_user,
        acao="FICHA_ADMISSIONAL_IMPORTADA",
        origem_importacao=True,
    )
    return {"mensagem": "Ficha admissional importada com sucesso.", "ficha": ficha}


def historico_salarial(db: Session, user_id: int, current_user: User) -> list[dict]:
    user = users_service.buscar_usuario(db, user_id)
    movimentos = historico_salarial_repository.listar_por_usuario(db, user_id)
    db.add(Log(
        user_id=current_user.id,
        acao="HISTORICO_SALARIAL_CONSULTADO",
        detalhes=f"Historico salarial de {user.nome} consultado por administrador",
    ))
    db.commit()
    return [
        {
            "data_vigencia": movimento.data_vigencia,
            "salario": _valor_decimal(descriptografar_dado_sensivel(movimento.salario_criptografado)),
            "tipo": movimento.tipo,
            "motivo": movimento.motivo,
            "criado_em": movimento.criado_em,
        }
        for movimento in movimentos
    ]


def historico_funcional(db: Session, user_id: int, current_user: User) -> list[dict]:
    user = users_service.buscar_usuario(db, user_id)
    resultado = historico_colaborador_service.listar_historico(db, user_id)
    db.add(Log(
        user_id=current_user.id,
        acao="HISTORICO_FUNCIONAL_CONSULTADO",
        detalhes=f"Historico funcional (cargo/departamento/beneficios) de {user.nome} consultado por administrador",
    ))
    db.commit()
    return resultado
