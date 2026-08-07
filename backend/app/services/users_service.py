from io import BytesIO
import json

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError

from app.core.security import hash_senha, verificar_senha
from app.core.timezone import hoje_sao_paulo
from app.core.cpf import formatar_cpf, mascarar_cpf, validar_cpf
from app.core.crypto import (
    criptografar_dado_sensivel,
    descriptografar_dado_sensivel,
    hash_dado_sensivel,
)
from app.models.log import Log
from app.models.user import User
from app.repositories import users_repository
from app.schemas.user import (
    ContatoEmergencia,
    ContatoEmergenciaOut,
    DadosBancarios,
    Endereco,
    UserConfigUpdate,
    UserCreate,
    UserUpdate,
)
from app.services import cargos_service, historico_colaborador_service, log_service


def calcular_dias_restantes(user: User, db: Session) -> int:
    from app.services.ferias_service import calcular_saldo

    return calcular_saldo(db, user)


def calcular_dias_usados(user: User, db: Session) -> int:
    from app.services.ferias_service import calcular_extrato_saldo

    return calcular_extrato_saldo(db, user)["dias_usados_total"]


def formatar_usuario(
    user: User,
    db: Session,
    dias_restantes: int | None = None,
    dias_usados_total: int | None = None,
) -> dict:
    departamento = None
    if user.departamento_id:
        dep = user.departamento
        departamento = {"id": dep.id, "nome": dep.nome} if dep else None

    return {
        "id": user.id,
        "nome": user.nome,
        "email": user.email,
        "role": user.role,
        "dias_totais": user.dias_totais,
        "dias_restantes": calcular_dias_restantes(user, db) if dias_restantes is None else dias_restantes,
        "dias_usados_total": calcular_dias_usados(user, db) if dias_usados_total is None else dias_usados_total,
        "departamento_id": user.departamento_id,
        "departamento": departamento,
        "data_admissao": user.data_admissao,
        "data_aniversario": user.data_aniversario,
        "cor": user.cor,
        "telefone": user.telefone,
        "cpf_mascarado": mascarar_cpf(descriptografar_dado_sensivel(user.cpf_criptografado)),
        "cargo": user.cargo.nome if user.cargo else None,
        "ativo": user.ativo,
        "must_change_password": getattr(user, "must_change_password", False),
        "saldo_manual_dias": user.saldo_manual_dias,
        "proxima_concessao_ferias": getattr(user, "proxima_concessao_ferias", None),
        "criado_em": user.criado_em,
    }


def formatar_dados_sensiveis(user: User) -> dict:
    return {
        "cpf": formatar_cpf(descriptografar_dado_sensivel(user.cpf_criptografado)) if user.cpf_criptografado else None,
        "contato_emergencia_1": _desserializar_contato_emergencia(user.contato_emergencia_1),
        "contato_emergencia_2": _desserializar_contato_emergencia(user.contato_emergencia_2),
        "endereco": _desserializar_endereco(user.endereco),
        "dados_bancarios": _desserializar_dados_bancarios(user.dados_bancarios),
    }


def exportar_usuarios_xlsx(db: Session, current_user: User) -> bytes:
    modelos = users_repository.listar_usuarios(db)
    usuarios = listar_usuarios(db, modelos)
    usuarios_model = {usuario.id: usuario for usuario in modelos}
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Colaboradores"

    cabecalhos = [
        "Nome",
        "E-mail",
        "CPF",
        "Cargo",
        "Departamento",
        "Telefone",
        "Contato de emergência 1 - Telefone",
        "Contato de emergência 1 - Nome",
        "Contato de emergência 1 - Grau de parentesco",
        "Contato de emergência 2 - Telefone",
        "Contato de emergência 2 - Nome",
        "Contato de emergência 2 - Grau de parentesco",
        "Perfil",
        "Status",
        "Data de admissão",
        "Data de aniversário",
        "Saldo de férias",
        "Dias usados",
        "Próxima concessão",
        "Endereço - Logradouro",
        "Endereço - Número",
        "Endereço - Bairro",
        "Endereço - Cidade",
        "Endereço - CEP",
        "Banco",
        "Agência",
        "Conta",
        "CPF do titular",
        "Nome do titular",
        "Chave PIX",
    ]
    planilha.append(cabecalhos)

    for usuario in usuarios:
        dados_sensiveis = formatar_dados_sensiveis(usuarios_model[usuario["id"]])
        endereco = dados_sensiveis["endereco"] or {}
        dados_bancarios = dados_sensiveis["dados_bancarios"] or {}
        contato_1 = dados_sensiveis["contato_emergencia_1"] or {}
        contato_2 = dados_sensiveis["contato_emergencia_2"] or {}
        planilha.append([
            usuario["nome"],
            usuario["email"],
            dados_sensiveis["cpf"] or "",
            usuario["cargo"] or "",
            (usuario["departamento"] or {}).get("nome", ""),
            usuario["telefone"] or "",
            contato_1.get("telefone", ""),
            contato_1.get("nome", ""),
            contato_1.get("grau_parentesco", ""),
            contato_2.get("telefone", ""),
            contato_2.get("nome", ""),
            contato_2.get("grau_parentesco", ""),
            "Administrador" if usuario["role"] == "admin" else "Usuário",
            "Ativo" if usuario["ativo"] else "Inativo",
            usuario["data_admissao"],
            usuario["data_aniversario"],
            usuario["dias_restantes"],
            usuario["dias_usados_total"],
            usuario["proxima_concessao_ferias"],
            endereco.get("logradouro", ""),
            endereco.get("numero", ""),
            endereco.get("bairro", ""),
            endereco.get("cidade", ""),
            endereco.get("cep", ""),
            dados_bancarios.get("banco", ""),
            dados_bancarios.get("agencia", ""),
            dados_bancarios.get("conta", ""),
            dados_bancarios.get("cpf_titular", ""),
            dados_bancarios.get("nome_titular", ""),
            dados_bancarios.get("chave_pix", ""),
        ])

    preenchimento = PatternFill("solid", fgColor="14213D")
    for celula in planilha[1]:
        celula.fill = preenchimento
        celula.font = Font(color="FFFFFF", bold=True)
        celula.alignment = Alignment(horizontal="center")

    for coluna in ("O", "P", "S"):
        for celula in planilha[coluna][1:]:
            if celula.value:
                celula.number_format = "dd/mm/yyyy"

    larguras = [
        30, 32, 18, 22, 24, 18, 18, 22, 20, 18, 22, 20, 16, 12, 18, 20, 18,
        14, 18, 32, 12, 22, 22, 14, 20, 14, 16, 18, 28, 30,
    ]
    for indice, largura in enumerate(larguras, start=1):
        planilha.column_dimensions[get_column_letter(indice)].width = largura

    planilha.freeze_panes = "A2"
    planilha.auto_filter.ref = planilha.dimensions

    arquivo = BytesIO()
    workbook.save(arquivo)

    log = log_service.construir_log(
        current_user,
        acao="DADOS_SENSIVEIS_USUARIOS_EXPORTADOS",
        detalhes=f"Planilha confidencial com dados de {len(usuarios)} usuário(s) exportada por administrador",
    )
    if log is not None:
        db.add(log)
    db.commit()
    return arquivo.getvalue()


def _serializar_endereco(endereco: Endereco | None) -> str | None:
    if not endereco:
        return None
    valores = endereco.model_dump(exclude_none=True)
    if not valores:
        return None
    conteudo = json.dumps(valores, ensure_ascii=False, separators=(",", ":"))
    return criptografar_dado_sensivel(conteudo)


def _desserializar_endereco(valor: str | None) -> dict | None:
    if not valor:
        return None
    valor = descriptografar_dado_sensivel(valor)
    try:
        dados = json.loads(valor)
        return Endereco.model_validate(dados).model_dump()
    except (json.JSONDecodeError, TypeError, ValueError):
        # Compatibilidade com registros anteriores, salvos como texto livre.
        return Endereco(logradouro=valor[:200]).model_dump()


def _serializar_contato_emergencia(contato: ContatoEmergencia | None) -> str | None:
    if not contato:
        return None
    valores = contato.model_dump(exclude_none=True)
    if not valores:
        return None
    conteudo = json.dumps(valores, ensure_ascii=False, separators=(",", ":"))
    return criptografar_dado_sensivel(conteudo)


def _desserializar_contato_emergencia(valor: str | None) -> dict | None:
    if not valor:
        return None
    valor = descriptografar_dado_sensivel(valor)
    try:
        dados = json.loads(valor)
        # Usa o formato tolerante para ler: cadastros migrados podem ter
        # nome/grau de parentesco em branco, e isso nao pode quebrar a leitura
        # (a obrigatoriedade dos 3 campos vale só para criar/editar).
        return ContatoEmergenciaOut.model_validate(dados).model_dump()
    except (json.JSONDecodeError, TypeError, ValueError):
        # Compatibilidade com registros legados: apenas o telefone, sem nome/parentesco.
        return {"telefone": valor, "nome": "", "grau_parentesco": ""}


def _serializar_dados_bancarios(dados: DadosBancarios | None) -> str | None:
    if not dados:
        return None
    valores = dados.model_dump(exclude_none=True)
    if not valores:
        return None
    conteudo = json.dumps(valores, ensure_ascii=False, separators=(",", ":"))
    return criptografar_dado_sensivel(conteudo)


def _desserializar_dados_bancarios(valor_criptografado: str | None) -> dict | None:
    valor = descriptografar_dado_sensivel(valor_criptografado)
    if not valor:
        return None
    try:
        dados = json.loads(valor)
        return DadosBancarios.model_validate(dados).model_dump()
    except (json.JSONDecodeError, TypeError, ValueError):
        # Compatibilidade com registros anteriores, salvos como texto livre.
        return DadosBancarios(banco=valor[:100]).model_dump()


def consultar_dados_sensiveis(db: Session, user_id: int, current_user: User) -> dict:
    user = buscar_usuario(db, user_id)
    log = log_service.construir_log(
        current_user,
        acao="CPF_COMPLETO_E_DADOS_SENSIVEIS_CONSULTADOS",
        detalhes=f"CPF completo e dados sensíveis do usuário #{user_id} consultados por administrador",
    )
    if log is not None:
        db.add(log)
    db.commit()
    return formatar_dados_sensiveis(user)


def buscar_usuario(db: Session, user_id: int) -> User:
    user = users_repository.obter_usuario_por_id(db, user_id)
    # O admin de sistema e tratado como inexistente para qualquer busca por
    # ID feita por outro usuario, fechando o acesso direto que contornaria a
    # ocultacao das listagens.
    if not user or user.is_sistema:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    return user


def validar_email_disponivel(db: Session, email: str, user_id: int | None = None) -> None:
    if user_id is None:
        existente = users_repository.obter_usuario_por_email(db, email)
        if existente:
            raise HTTPException(status_code=400, detail="E-mail já cadastrado")
        return

    existente = users_repository.obter_usuario_por_email_exceto_id(db, email, user_id)
    if existente:
        raise HTTPException(status_code=400, detail="E-mail já em uso por outro usuário")


def validar_departamento(db: Session, departamento_id: int | None) -> None:
    if departamento_id and not users_repository.obter_departamento_por_id(db, departamento_id):
        raise HTTPException(status_code=404, detail="Departamento não encontrado")


def preparar_cpf(db: Session, cpf: str, excluir_user_id: int | None = None) -> tuple[str, str]:
    try:
        normalizado = validar_cpf(cpf)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="CPF inválido") from exc
    cpf_hash = hash_dado_sensivel(normalizado)
    if users_repository.obter_usuario_por_cpf_hash(db, cpf_hash, excluir_user_id):
        raise HTTPException(status_code=400, detail="CPF já cadastrado para outro colaborador")
    return criptografar_dado_sensivel(normalizado), cpf_hash


def listar_usuarios(db: Session, users: list[User] | None = None) -> list[dict]:
    users = users if users is not None else users_repository.listar_usuarios(db)
    resumos = users_repository.resumir_saldos_usuarios(db, [user.id for user in users])
    resultado = []
    for user in users:
        extrato = resumos.get(user.id, {"saldo": 0, "dias_usados_total": 0})
        resultado.append(
            formatar_usuario(
                user,
                db,
                dias_restantes=extrato["saldo"],
                dias_usados_total=extrato["dias_usados_total"],
            )
        )
    return resultado


def criar_usuario(
    db: Session,
    payload: UserCreate,
    current_user: User,
    commit: bool = True,
) -> dict:
    validar_email_disponivel(db, payload.email)
    validar_departamento(db, payload.departamento_id)
    cargo = cargos_service.obter_cargo_por_nome(db, payload.cargo)
    cpf_criptografado, cpf_hash = preparar_cpf(db, payload.cpf)

    novo_user = User(
        nome=payload.nome,
        email=payload.email,
        senha_hash=hash_senha(payload.senha),
        role=payload.role,
        dias_totais=payload.dias_totais,
        proxima_concessao_ferias=payload.proxima_concessao_ferias,
        departamento_id=payload.departamento_id,
        data_admissao=payload.data_admissao,
        data_aniversario=payload.data_aniversario,
        cor=payload.cor,
        telefone=payload.telefone,
        contato_emergencia_1=_serializar_contato_emergencia(payload.contato_emergencia_1),
        contato_emergencia_2=_serializar_contato_emergencia(payload.contato_emergencia_2),
        endereco=_serializar_endereco(payload.endereco),
        dados_bancarios=_serializar_dados_bancarios(payload.dados_bancarios),
        cargo_id=cargo.id if cargo else None,
        cpf_criptografado=cpf_criptografado,
        cpf_hash=cpf_hash,
        must_change_password=True,
    )

    log = (
        Log(
            user_id=novo_user.id,
            acao="USUARIO_CRIADO",
            detalhes=f"Usuário {novo_user.nome} ({novo_user.email}) criado por {current_user.nome}",
        )
        if not getattr(current_user, "is_sistema", False)
        else None
    )
    try:
        users_repository.salvar_usuario_com_log(db, novo_user, log, commit=False)

        if cargo:
            historico_colaborador_service.processar_alteracao_se_necessario(
                db, novo_user.id, "cargo", None, cargo.nome, None, None, current_user,
            )
        if novo_user.departamento_id:
            departamento = users_repository.obter_departamento_por_id(db, novo_user.departamento_id)
            if departamento:
                historico_colaborador_service.processar_alteracao_se_necessario(
                    db, novo_user.id, "departamento", None, departamento.nome, None, None, current_user,
                )

        from app.services.ferias_service import registrar_saldo_inicial

        registrar_saldo_inicial(
            db,
            novo_user,
            payload.saldo_inicial_dias,
            current_user.id,
            commit=False,
        )
        if commit:
            db.commit()
            db.refresh(novo_user)
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="E-mail ou CPF já cadastrado") from exc

    return formatar_usuario(novo_user, db)


CAMPOS_DIFF_USUARIO = (
    ("nome", "nome"),
    ("email", "e-mail"),
    ("dias_totais", "dias de férias por ano"),
    ("proxima_concessao_ferias", "próxima concessão de férias"),
    ("data_admissao", "data de admissão"),
    ("data_aniversario", "data de aniversário"),
    ("cor", "cor"),
    ("telefone", "telefone"),
)


def editar_usuario(db: Session, user_id: int, payload: UserUpdate, current_user: User) -> dict:
    user = buscar_usuario(db, user_id)
    cargo_anterior_nome = user.cargo.nome if user.cargo else None
    departamento_anterior_nome = user.departamento.nome if user.departamento else None
    departamento_novo_nome = departamento_anterior_nome
    valores_anteriores = {campo: getattr(user, campo, None) for campo, _ in CAMPOS_DIFF_USUARIO}
    # A criptografia usada não é determinística (mesmo texto gera cifra
    # diferente a cada vez), então comparar o valor cru já criptografado
    # sempre acusaria mudança. Descriptografa antes/depois para comparar o
    # conteúdo de verdade — o formulário reenvia esses campos preenchidos
    # em todo submit, mesmo quando o admin não mexeu neles.
    cpf_armazenado = getattr(user, "cpf_criptografado", None)
    cpf_anterior = descriptografar_dado_sensivel(cpf_armazenado) if cpf_armazenado else None
    contato_emergencia_1_anterior = _desserializar_contato_emergencia(
        getattr(user, "contato_emergencia_1", None)
    )
    contato_emergencia_2_anterior = _desserializar_contato_emergencia(
        getattr(user, "contato_emergencia_2", None)
    )
    endereco_anterior = _desserializar_endereco(getattr(user, "endereco", None))
    dados_bancarios_anterior = _desserializar_dados_bancarios(
        getattr(user, "dados_bancarios", None)
    )

    if payload.nome is not None:
        user.nome = payload.nome
    if payload.email is not None:
        validar_email_disponivel(db, payload.email, user_id)
        user.email = payload.email
    if payload.dias_totais is not None:
        user.dias_totais = payload.dias_totais
    if "proxima_concessao_ferias" in payload.model_fields_set:
        user.proxima_concessao_ferias = payload.proxima_concessao_ferias
    if payload.departamento_id is not None:
        if payload.departamento_id == 0:
            user.departamento_id = None
            departamento_novo_nome = None
        else:
            validar_departamento(db, payload.departamento_id)
            user.departamento_id = payload.departamento_id
            novo_departamento = users_repository.obter_departamento_por_id(db, payload.departamento_id)
            departamento_novo_nome = novo_departamento.nome if novo_departamento else None
    if payload.data_admissao is not None:
        user.data_admissao = payload.data_admissao
    if payload.data_aniversario is not None:
        user.data_aniversario = payload.data_aniversario
    if payload.senha is not None and payload.senha.strip():
        user.senha_hash = hash_senha(payload.senha)
        user.token_version += 1
        user.must_change_password = True
    if payload.cor is not None:
        user.cor = payload.cor if payload.cor.strip() else None
    if "telefone" in payload.model_fields_set:
        user.telefone = payload.telefone.strip() if payload.telefone and payload.telefone.strip() else None
    if "contato_emergencia_1" in payload.model_fields_set:
        user.contato_emergencia_1 = _serializar_contato_emergencia(payload.contato_emergencia_1)
    if "contato_emergencia_2" in payload.model_fields_set:
        user.contato_emergencia_2 = _serializar_contato_emergencia(payload.contato_emergencia_2)
    if "endereco" in payload.model_fields_set:
        user.endereco = _serializar_endereco(payload.endereco)
    if "dados_bancarios" in payload.model_fields_set:
        user.dados_bancarios = _serializar_dados_bancarios(payload.dados_bancarios)
    cargo_novo_nome = cargo_anterior_nome
    if "cargo" in payload.model_fields_set:
        cargo = cargos_service.obter_cargo_por_nome(db, payload.cargo)
        user.cargo_id = cargo.id if cargo else None
        cargo_novo_nome = cargo.nome if cargo else None
    if "cpf" in payload.model_fields_set and payload.cpf:
        user.cpf_criptografado, user.cpf_hash = preparar_cpf(db, payload.cpf, user.id)

    historico_colaborador_service.processar_alteracao_se_necessario(
        db, user.id, "cargo", cargo_anterior_nome, cargo_novo_nome,
        payload.motivo_alteracao_funcional, payload.tipo_alteracao_funcional, current_user,
    )
    historico_colaborador_service.processar_alteracao_se_necessario(
        db, user.id, "departamento", departamento_anterior_nome, departamento_novo_nome,
        payload.motivo_alteracao_funcional, payload.tipo_alteracao_funcional, current_user,
    )

    if payload.saldo_atual_dias is not None:
        from app.services.ferias_service import (
            ajustar_saldo,
            calcular_saldo,
            garantir_saldo_atualizado,
        )
        garantir_saldo_atualizado(db, user, current_user.id, commit=False)
        if payload.saldo_atual_dias != calcular_saldo(db, user):
            if not payload.motivo_ajuste_saldo:
                raise HTTPException(status_code=400, detail="Informe o motivo do ajuste de saldo")
            ajustar_saldo(
                db,
                user,
                payload.saldo_atual_dias,
                payload.motivo_ajuste_saldo,
                current_user,
                commit=False,
            )

    alteracoes = []
    for campo, rotulo in CAMPOS_DIFF_USUARIO:
        anterior = valores_anteriores[campo]
        novo = getattr(user, campo, None)
        if novo != anterior:
            alteracoes.append(f"{rotulo}: {anterior if anterior is not None else '—'} -> {novo if novo is not None else '—'}")
    if cargo_novo_nome != cargo_anterior_nome:
        alteracoes.append(f"cargo: {cargo_anterior_nome or '—'} -> {cargo_novo_nome or '—'}")
    if departamento_novo_nome != departamento_anterior_nome:
        alteracoes.append(f"departamento: {departamento_anterior_nome or '—'} -> {departamento_novo_nome or '—'}")
    if payload.senha is not None and payload.senha.strip():
        alteracoes.append("senha redefinida")

    cpf_armazenado_novo = getattr(user, "cpf_criptografado", None)
    cpf_novo = descriptografar_dado_sensivel(cpf_armazenado_novo) if cpf_armazenado_novo else None
    contato_emergencia_1_novo = _desserializar_contato_emergencia(
        getattr(user, "contato_emergencia_1", None)
    )
    contato_emergencia_2_novo = _desserializar_contato_emergencia(
        getattr(user, "contato_emergencia_2", None)
    )
    endereco_novo = _desserializar_endereco(getattr(user, "endereco", None))
    dados_bancarios_novo = _desserializar_dados_bancarios(
        getattr(user, "dados_bancarios", None)
    )

    # Valores de campos sensiveis/criptografados (senha, contatos de
    # emergencia, endereco, dados bancarios, CPF) nao aparecem no log — só o
    # nome do campo — para nao expor em texto plano dados que sao
    # armazenados criptografados justamente por serem sensiveis.
    if contato_emergencia_1_novo != contato_emergencia_1_anterior:
        alteracoes.append("contato de emergência 1 alterado")
    if contato_emergencia_2_novo != contato_emergencia_2_anterior:
        alteracoes.append("contato de emergência 2 alterado")
    if endereco_novo != endereco_anterior:
        alteracoes.append("endereço alterado")
    if dados_bancarios_novo != dados_bancarios_anterior:
        alteracoes.append("dados bancários alterados")
    if cpf_novo != cpf_anterior:
        alteracoes.append("CPF alterado")
    resumo = "; ".join(alteracoes) if alteracoes else "nenhum campo alterado"

    log = log_service.construir_log(
        current_user,
        acao="USUARIO_EDITADO",
        detalhes=f"Usuário {user.nome} ({user.email}) editado por {current_user.nome}: {resumo}",
    )
    try:
        users_repository.atualizar_usuario_com_log(db, user, log)
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="E-mail ou CPF já cadastrado") from exc

    return formatar_usuario(user, db)


def listar_aniversariantes(db: Session) -> list[dict]:
    hoje = hoje_sao_paulo()
    return [
        {
            "nome": user.nome,
            "dia": user.data_aniversario.day,
            "mes": user.data_aniversario.month,
        }
        for user in users_repository.listar_usuarios_com_aniversario(db)
        if user.data_aniversario and user.data_aniversario.month == hoje.month
    ]


def desativar_usuario(db: Session, user_id: int, current_user: User) -> None:
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Você não pode excluir sua própria conta")

    user = buscar_usuario(db, user_id)
    if user.role == "admin" and users_repository.contar_administradores_ativos(db) <= 1:
        raise HTTPException(status_code=400, detail="O último administrador ativo não pode ser desativado")
    if not user.ativo:
        return
    from app.repositories import patrimonios_repository

    if patrimonios_repository.existe_fluxo_aberto_para_usuario(db, user.id):
        raise HTTPException(
            status_code=409,
            detail=(
                "O colaborador possui autorização de equipamento em aberto. "
                "Conclua ou encerre o fluxo antes de desativar a conta."
            ),
        )
    user.ativo = False
    log = log_service.construir_log(
        current_user,
        acao="USUARIO_DESATIVADO",
        detalhes=f"Usuário {user.nome} ({user.email}) desativado por {current_user.nome}",
    )
    try:
        users_repository.atualizar_usuario_com_log(db, user, log)
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="E-mail ou CPF já cadastrado") from exc


def reativar_usuario(db: Session, user_id: int, current_user: User) -> dict:
    user = buscar_usuario(db, user_id)
    if user.ativo:
        return formatar_usuario(user, db)
    user.ativo = True
    log = log_service.construir_log(
        current_user,
        acao="USUARIO_REATIVADO",
        detalhes=f"Usuário {user.nome} ({user.email}) reativado por {current_user.nome}",
    )
    users_repository.atualizar_usuario_com_log(db, user, log)
    return formatar_usuario(user, db)


def atualizar_configuracoes(db: Session, payload: UserConfigUpdate, current_user: User) -> dict:
    troca_obrigatoria = getattr(current_user, "must_change_password", False)
    senha_alterada = bool(payload.nova_senha and payload.nova_senha.strip())
    if troca_obrigatoria and not senha_alterada:
        raise HTTPException(
            status_code=400,
            detail="A troca da senha temporária é obrigatória antes de continuar",
        )
    if payload.nome is not None:
        current_user.nome = payload.nome
    if payload.email is not None:
        validar_email_disponivel(db, payload.email, current_user.id)
        current_user.email = payload.email
    if "telefone" in payload.model_fields_set:
        current_user.telefone = payload.telefone
    if "contato_emergencia_1" in payload.model_fields_set:
        current_user.contato_emergencia_1 = _serializar_contato_emergencia(payload.contato_emergencia_1)
    if "contato_emergencia_2" in payload.model_fields_set:
        current_user.contato_emergencia_2 = _serializar_contato_emergencia(payload.contato_emergencia_2)

    if senha_alterada:
        if not payload.senha_atual:
            raise HTTPException(status_code=400, detail="Informe a senha atual para alterá-la")
        if not verificar_senha(payload.senha_atual, current_user.senha_hash):
            raise HTTPException(status_code=400, detail="Senha atual incorreta")
        if verificar_senha(payload.nova_senha, current_user.senha_hash):
            raise HTTPException(status_code=400, detail="A nova senha deve ser diferente da senha atual")
        current_user.senha_hash = hash_senha(payload.nova_senha)
        current_user.token_version += 1
        current_user.must_change_password = False

    if troca_obrigatoria:
        acao = "SENHA_TEMPORARIA_SUBSTITUIDA"
        detalhes = "Senha temporária substituída no primeiro acesso"
    elif senha_alterada:
        acao = "SENHA_ALTERADA"
        detalhes = "Senha da conta alterada"
    else:
        acao = "PERFIL_ATUALIZADO"
        detalhes = "Dados do perfil atualizados"

    log = log_service.construir_log(current_user, acao=acao, detalhes=detalhes)
    if log is not None:
        db.add(log)
    try:
        users_repository.salvar_usuario(db, current_user)
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="E-mail já cadastrado") from exc
    return formatar_usuario(current_user, db)


def meu_perfil(db: Session, current_user: User) -> dict:
    dados_sensiveis = formatar_dados_sensiveis(current_user)
    base = formatar_usuario(current_user, db)
    return {
        "id": base["id"],
        "nome": base["nome"],
        "email": base["email"],
        "role": base["role"],
        "cor": base["cor"],
        "cargo": base["cargo"],
        "departamento": base["departamento"],
        "data_admissao": base["data_admissao"],
        "data_aniversario": base["data_aniversario"],
        "telefone": base["telefone"],
        "cpf": dados_sensiveis["cpf"],
        "contato_emergencia_1": dados_sensiveis["contato_emergencia_1"],
        "contato_emergencia_2": dados_sensiveis["contato_emergencia_2"],
        "endereco": dados_sensiveis["endereco"],
        "dados_bancarios": dados_sensiveis["dados_bancarios"],
    }
