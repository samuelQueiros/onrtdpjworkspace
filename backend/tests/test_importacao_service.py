from datetime import date, datetime
import unittest
from types import SimpleNamespace
from unittest.mock import patch

from fastapi import HTTPException
from openpyxl import load_workbook

from app.services import importacao_service


class ImportacaoServiceTests(unittest.TestCase):
    def test_parse_date_aceita_formatos_suportados(self):
        self.assertEqual(importacao_service.parse_date("2026-07-05"), date(2026, 7, 5))
        self.assertEqual(importacao_service.parse_date("05/07/2026"), date(2026, 7, 5))
        self.assertEqual(importacao_service.parse_date("05-07-2026"), date(2026, 7, 5))

    def test_validar_extensao_planilha_rejeita_extensao_invalida(self):
        with self.assertRaises(HTTPException) as exc:
            importacao_service.validar_extensao_planilha("arquivo.csv")

        self.assertEqual(exc.exception.status_code, 400)

    def test_importar_ferias_insere_registro_e_log(self):
        rows = [
            ("email", "data_inicio", "data_fim", "ferias_acordo"),
            ("gabriel@sistema.com", "2027-07-05", "2027-07-09", False),
        ]
        user = SimpleNamespace(id=2)
        current_user = SimpleNamespace(id=1)

        with (
            patch("app.services.importacao_service.carregar_linhas_planilha", return_value=rows),
            patch("app.services.importacao_service.importacao_repository.obter_usuario_por_email", return_value=user),
            patch("app.services.importacao_service.importacao_repository.existe_ferias_periodo", return_value=False),
            patch("app.services.importacao_service.ferias_service.bloquear_regras_ferias"),
            patch("app.services.importacao_service.ferias_service.garantir_saldo_atualizado"),
            patch("app.services.importacao_service.ferias_service.verificar_regras_data"),
            patch("app.services.importacao_service.ferias_service.verificar_bloqueio_datas"),
            patch("app.services.importacao_service.ferias_service.verificar_sobreposicao_usuario"),
            patch("app.services.importacao_service.ferias_service.verificar_sobreposicao_departamento", return_value=False),
            patch("app.services.importacao_service.ferias_service.calcular_saldo", return_value=30),
            patch("app.services.importacao_service.importacao_repository.adicionar_ferias") as adicionar_ferias,
            patch("app.services.importacao_service.importacao_repository.adicionar_log") as adicionar_log,
            patch("app.services.importacao_service.importacao_repository.commit") as commit,
        ):
            response = importacao_service.importar_ferias(SimpleNamespace(), "ferias.xlsx", b"conteudo", current_user)

        self.assertEqual(response["inseridos"], 1)
        adicionar_ferias.assert_called_once()
        adicionar_log.assert_called_once()
        commit.assert_called_once()

    def test_importar_ferias_registra_erro_para_usuario_inexistente(self):
        rows = [("email", "data_inicio", "data_fim"), ("naoexiste@sistema.com", "2027-07-05", "2027-07-09")]

        with (
            patch("app.services.importacao_service.carregar_linhas_planilha", return_value=rows),
            patch("app.services.importacao_service.importacao_repository.obter_usuario_por_email", return_value=None),
            patch("app.services.importacao_service.importacao_repository.commit") as commit,
        ):
            response = importacao_service.importar_ferias(SimpleNamespace(), "ferias.xlsx", b"conteudo", SimpleNamespace(id=1))

        self.assertEqual(response["inseridos"], 0)
        self.assertEqual(len(response["erros"]), 1)
        commit.assert_not_called()

    def test_importar_logs_insere_log_com_data_parseada(self):
        rows = [("data", "acao", "detalhes"), ("05/07/2026 10:30", "ACAO", "Detalhe")]

        with (
            patch("app.services.importacao_service.carregar_linhas_planilha", return_value=rows),
            patch("app.services.importacao_service.importacao_repository.adicionar_log") as adicionar_log,
            patch("app.services.importacao_service.importacao_repository.commit") as commit,
        ):
            response = importacao_service.importar_logs(SimpleNamespace(), "logs.xlsx", b"conteudo", SimpleNamespace(id=1))

        self.assertEqual(response["inseridos"], 1)
        log = adicionar_log.call_args_list[0].args[1]
        self.assertEqual(log.criado_em, datetime(2026, 7, 5, 10, 30))
        self.assertEqual(adicionar_log.call_count, 2)
        commit.assert_called_once()

    def test_gerar_modelo_colaboradores_inclui_abas_e_campos(self):
        conteudo = importacao_service.gerar_modelo_colaboradores_xlsx()
        workbook = load_workbook(importacao_service.io.BytesIO(conteudo))

        self.assertEqual(workbook.sheetnames, ["Instruções", "Colaboradores"])
        planilha = workbook["Colaboradores"]
        self.assertEqual(planilha["A1"].value, "Nome")
        self.assertEqual(planilha["AE1"].value, "Dias de férias por período")
        self.assertEqual(planilha.freeze_panes, "A2")

    def test_importar_colaboradores_valida_tudo_antes_de_inserir(self):
        rows = [
            tuple(importacao_service.CABECALHOS_COLABORADORES),
            (
                "Pessoa A", "pessoa@empresa.com", "529.982.247-25", "Analista",
                "Tecnologia", "61999999999",
                "61988888888", "Maria Exemplo", "Mãe",
                "61977777777", "João Exemplo", "Pai",
                "Usuário", "Ativo", "10/01/2024", "20/05/1995", 30, 0,
                "10/01/2027", "Rua A", "10", "Centro", "Brasília", "70000-000",
                "Banco", "0001", "123-4", "529.982.247-25", "Pessoa A",
                "pessoa@empresa.com", 30, "Temporaria@Pessoa1",
            ),
        ]
        departamento = SimpleNamespace(id=4, nome="Tecnologia")
        cargo = SimpleNamespace(id=5, nome="Analista")
        db = SimpleNamespace(add=lambda _: None, commit=lambda: None, rollback=lambda: None)
        current_user = SimpleNamespace(id=1, nome="Admin")

        with (
            patch("app.services.importacao_service.carregar_linhas_planilha", return_value=rows),
            patch("app.services.importacao_service.users_repository.listar_usuarios", return_value=[]),
            patch("app.services.importacao_service.users_repository.obter_usuario_por_email", return_value=None),
            patch("app.services.importacao_service.users_repository.obter_usuario_por_cpf_hash", return_value=None),
            patch("app.services.importacao_service.departamentos_repository.obter_departamento_por_nome", return_value=departamento),
            patch("app.services.importacao_service.cargos_repository.obter_cargo_por_nome", return_value=cargo),
            patch("app.services.importacao_service.users_service.criar_usuario") as criar,
        ):
            resultado = importacao_service.importar_colaboradores(
                db, "colaboradores.xlsx", b"conteudo", current_user
            )

        self.assertEqual(resultado["inseridos"], 1)
        payload = criar.call_args.args[1]
        self.assertEqual(payload.senha, "Temporaria@Pessoa1")
        self.assertIn(payload.cor, importacao_service.CORES_IMPORTACAO)
        criar.assert_called_once_with(db, payload, current_user, commit=False)

    def test_importar_colaboradores_nao_insere_quando_uma_linha_tem_erro(self):
        rows = [
            tuple(importacao_service.CABECALHOS_COLABORADORES),
            tuple([""] * len(importacao_service.CABECALHOS_COLABORADORES)),
            ("Pessoa sem dados",) + tuple([""] * (len(importacao_service.CABECALHOS_COLABORADORES) - 1)),
        ]
        db = SimpleNamespace()

        with (
            patch("app.services.importacao_service.carregar_linhas_planilha", return_value=rows),
            patch("app.services.importacao_service.users_repository.listar_usuarios", return_value=[]),
            patch("app.services.importacao_service.users_service.criar_usuario") as criar,
        ):
            resultado = importacao_service.importar_colaboradores(
                db, "colaboradores.xlsx", b"conteudo", SimpleNamespace(id=1)
            )

        self.assertEqual(resultado["inseridos"], 0)
        self.assertTrue(resultado["erros"])
        criar.assert_not_called()


if __name__ == "__main__":
    unittest.main()
