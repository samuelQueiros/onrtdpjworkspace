import unittest
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

from fastapi import HTTPException
from openpyxl import load_workbook

from app.models.ficha_admissional import FichaAdmissional
from app.models.historico_salarial import HistoricoSalarial
from app.schemas.ficha_admissional import FichaAdmissionalUpdate
from app.services import fichas_admissionais_service


class FakeDb:
    def __init__(self):
        self.added = []
        self.commits = 0

    def add(self, item):
        self.added.append(item)

    def commit(self):
        self.commits += 1

    def refresh(self, _item):
        pass


def ficha_vazia() -> FichaAdmissional:
    ficha = FichaAdmissional(user_id=7, status="rascunho")
    ficha.id = 3
    ficha.criado_por_id = 1
    ficha.atualizado_por_id = 1
    ficha.criado_em = datetime(2026, 7, 29, tzinfo=UTC)
    ficha.atualizado_em = datetime(2026, 7, 29, tzinfo=UTC)
    return ficha


class FichasAdmissionaisServiceTests(unittest.TestCase):
    def test_atualizar_ficha_criptografa_dados_sensiveis_e_audita(self):
        db = FakeDb()
        ficha = ficha_vazia()
        payload = FichaAdmissionalUpdate(
            local_nascimento="Brasília",
            salario=Decimal("7500.50"),
            uf_nascimento="df",
            status="completa",
        )

        def descriptografar(valor):
            return valor.removeprefix("enc:") if valor else None

        with (
            patch("app.services.fichas_admissionais_service.users_service.buscar_usuario"),
            patch(
                "app.services.fichas_admissionais_service.fichas_admissionais_repository.obter_por_usuario",
                return_value=ficha,
            ),
            patch("app.services.fichas_admissionais_service.fichas_admissionais_repository.salvar"),
            patch(
                "app.services.fichas_admissionais_service.criptografar_dado_sensivel",
                side_effect=lambda valor: f"enc:{valor}",
            ),
            patch(
                "app.services.fichas_admissionais_service.descriptografar_dado_sensivel",
                side_effect=descriptografar,
            ),
        ):
            resultado = fichas_admissionais_service.atualizar_ficha(
                db,
                7,
                payload,
                SimpleNamespace(id=1),
            )

        self.assertEqual(ficha.local_nascimento_criptografado, "enc:Brasília")
        self.assertEqual(ficha.salario_criptografado, "enc:7500.50")
        self.assertEqual(ficha.uf_nascimento, "DF")
        self.assertEqual(resultado["salario"], Decimal("7500.50"))
        self.assertEqual(db.added[0].acao, "FICHA_ADMISSIONAL_ATUALIZADA")
        self.assertEqual(db.commits, 1)

    def test_atualizar_ficha_grava_historico_no_cadastro_inicial_de_salario(self):
        db = FakeDb()
        ficha = ficha_vazia()  # sem salario_criptografado definido
        payload = FichaAdmissionalUpdate(salario=Decimal("5000.00"))

        with (
            patch("app.services.fichas_admissionais_service.users_service.buscar_usuario"),
            patch(
                "app.services.fichas_admissionais_service.fichas_admissionais_repository.obter_por_usuario",
                return_value=ficha,
            ),
            patch("app.services.fichas_admissionais_service.fichas_admissionais_repository.salvar"),
            patch(
                "app.services.fichas_admissionais_service.criptografar_dado_sensivel",
                side_effect=lambda valor: f"enc:{valor}",
            ),
            patch(
                "app.services.fichas_admissionais_service.descriptografar_dado_sensivel",
                side_effect=lambda valor: valor.removeprefix("enc:") if valor else None,
            ),
        ):
            fichas_admissionais_service.atualizar_ficha(db, 7, payload, SimpleNamespace(id=1))

        historicos = [item for item in db.added if isinstance(item, HistoricoSalarial)]
        self.assertEqual(len(historicos), 1)
        self.assertEqual(historicos[0].motivo, "Cadastro inicial")
        self.assertEqual(historicos[0].salario_criptografado, "enc:5000.00")

    def test_atualizar_ficha_exige_motivo_ao_alterar_salario_existente(self):
        db = FakeDb()
        ficha = ficha_vazia()
        ficha.salario_criptografado = "enc:5000.00"
        payload = FichaAdmissionalUpdate(salario=Decimal("6000.00"))

        with (
            patch("app.services.fichas_admissionais_service.users_service.buscar_usuario"),
            patch(
                "app.services.fichas_admissionais_service.fichas_admissionais_repository.obter_por_usuario",
                return_value=ficha,
            ),
            patch("app.services.fichas_admissionais_service.fichas_admissionais_repository.salvar"),
            patch(
                "app.services.fichas_admissionais_service.criptografar_dado_sensivel",
                side_effect=lambda valor: f"enc:{valor}",
            ),
            patch(
                "app.services.fichas_admissionais_service.descriptografar_dado_sensivel",
                side_effect=lambda valor: valor.removeprefix("enc:") if valor else None,
            ),
        ):
            with self.assertRaises(HTTPException) as exc:
                fichas_admissionais_service.atualizar_ficha(db, 7, payload, SimpleNamespace(id=1))

        self.assertEqual(exc.exception.status_code, 400)
        self.assertEqual(db.commits, 0)

    def test_atualizar_ficha_grava_historico_com_motivo_de_reajuste(self):
        db = FakeDb()
        ficha = ficha_vazia()
        ficha.salario_criptografado = "enc:5000.00"
        payload = FichaAdmissionalUpdate(
            salario=Decimal("6000.00"),
            motivo_alteracao_salario="Promoção",
            tipo_alteracao_salario="reajuste",
        )

        with (
            patch("app.services.fichas_admissionais_service.users_service.buscar_usuario"),
            patch(
                "app.services.fichas_admissionais_service.fichas_admissionais_repository.obter_por_usuario",
                return_value=ficha,
            ),
            patch("app.services.fichas_admissionais_service.fichas_admissionais_repository.salvar"),
            patch(
                "app.services.fichas_admissionais_service.criptografar_dado_sensivel",
                side_effect=lambda valor: f"enc:{valor}",
            ),
            patch(
                "app.services.fichas_admissionais_service.descriptografar_dado_sensivel",
                side_effect=lambda valor: valor.removeprefix("enc:") if valor else None,
            ),
        ):
            fichas_admissionais_service.atualizar_ficha(db, 7, payload, SimpleNamespace(id=1))

        historicos = [item for item in db.added if isinstance(item, HistoricoSalarial)]
        self.assertEqual(len(historicos), 1)
        self.assertEqual(historicos[0].tipo, "reajuste")
        self.assertEqual(historicos[0].motivo, "Promoção")
        self.assertEqual(historicos[0].salario_criptografado, "enc:6000.00")

    def test_atualizar_ficha_grava_historico_como_correcao_quando_informado(self):
        db = FakeDb()
        ficha = ficha_vazia()
        ficha.salario_criptografado = "enc:5000.00"
        payload = FichaAdmissionalUpdate(
            salario=Decimal("4500.00"),
            motivo_alteracao_salario="Erro de digitação no cadastro anterior",
            tipo_alteracao_salario="correcao",
        )

        with (
            patch("app.services.fichas_admissionais_service.users_service.buscar_usuario"),
            patch(
                "app.services.fichas_admissionais_service.fichas_admissionais_repository.obter_por_usuario",
                return_value=ficha,
            ),
            patch("app.services.fichas_admissionais_service.fichas_admissionais_repository.salvar"),
            patch(
                "app.services.fichas_admissionais_service.criptografar_dado_sensivel",
                side_effect=lambda valor: f"enc:{valor}",
            ),
            patch(
                "app.services.fichas_admissionais_service.descriptografar_dado_sensivel",
                side_effect=lambda valor: valor.removeprefix("enc:") if valor else None,
            ),
        ):
            fichas_admissionais_service.atualizar_ficha(db, 7, payload, SimpleNamespace(id=1))

        historicos = [item for item in db.added if isinstance(item, HistoricoSalarial)]
        self.assertEqual(len(historicos), 1)
        self.assertEqual(historicos[0].tipo, "correcao")

    def test_atualizar_ficha_exige_tipo_alem_do_motivo(self):
        db = FakeDb()
        ficha = ficha_vazia()
        ficha.salario_criptografado = "enc:5000.00"
        payload = FichaAdmissionalUpdate(salario=Decimal("6000.00"), motivo_alteracao_salario="Promoção")

        with (
            patch("app.services.fichas_admissionais_service.users_service.buscar_usuario"),
            patch(
                "app.services.fichas_admissionais_service.fichas_admissionais_repository.obter_por_usuario",
                return_value=ficha,
            ),
            patch("app.services.fichas_admissionais_service.fichas_admissionais_repository.salvar"),
            patch(
                "app.services.fichas_admissionais_service.criptografar_dado_sensivel",
                side_effect=lambda valor: f"enc:{valor}",
            ),
            patch(
                "app.services.fichas_admissionais_service.descriptografar_dado_sensivel",
                side_effect=lambda valor: valor.removeprefix("enc:") if valor else None,
            ),
        ):
            with self.assertRaises(HTTPException) as exc:
                fichas_admissionais_service.atualizar_ficha(db, 7, payload, SimpleNamespace(id=1))

        self.assertEqual(exc.exception.status_code, 400)

    def test_atualizar_ficha_importacao_nao_exige_motivo(self):
        db = FakeDb()
        ficha = ficha_vazia()
        ficha.salario_criptografado = "enc:5000.00"
        payload = FichaAdmissionalUpdate(salario=Decimal("6000.00"))

        with (
            patch("app.services.fichas_admissionais_service.users_service.buscar_usuario"),
            patch(
                "app.services.fichas_admissionais_service.fichas_admissionais_repository.obter_por_usuario",
                return_value=ficha,
            ),
            patch("app.services.fichas_admissionais_service.fichas_admissionais_repository.salvar"),
            patch(
                "app.services.fichas_admissionais_service.criptografar_dado_sensivel",
                side_effect=lambda valor: f"enc:{valor}",
            ),
            patch(
                "app.services.fichas_admissionais_service.descriptografar_dado_sensivel",
                side_effect=lambda valor: valor.removeprefix("enc:") if valor else None,
            ),
        ):
            fichas_admissionais_service.atualizar_ficha(
                db, 7, payload, SimpleNamespace(id=1), origem_importacao=True,
            )

        historicos = [item for item in db.added if isinstance(item, HistoricoSalarial)]
        self.assertEqual(len(historicos), 1)
        self.assertEqual(historicos[0].motivo, "Importação de planilha")

    def test_atualizar_ficha_nao_grava_historico_quando_salario_nao_muda(self):
        db = FakeDb()
        ficha = ficha_vazia()
        ficha.salario_criptografado = "enc:5000.00"
        payload = FichaAdmissionalUpdate(salario=Decimal("5000.00"))

        with (
            patch("app.services.fichas_admissionais_service.users_service.buscar_usuario"),
            patch(
                "app.services.fichas_admissionais_service.fichas_admissionais_repository.obter_por_usuario",
                return_value=ficha,
            ),
            patch("app.services.fichas_admissionais_service.fichas_admissionais_repository.salvar"),
            patch(
                "app.services.fichas_admissionais_service.criptografar_dado_sensivel",
                side_effect=lambda valor: f"enc:{valor}",
            ),
            patch(
                "app.services.fichas_admissionais_service.descriptografar_dado_sensivel",
                side_effect=lambda valor: valor.removeprefix("enc:") if valor else None,
            ),
        ):
            fichas_admissionais_service.atualizar_ficha(db, 7, payload, SimpleNamespace(id=1))

        historicos = [item for item in db.added if isinstance(item, HistoricoSalarial)]
        self.assertEqual(len(historicos), 0)

    def test_historico_salarial_descriptografa_e_audita(self):
        db = FakeDb()
        movimentos = [
            SimpleNamespace(
                data_vigencia=date(2026, 1, 1),
                salario_criptografado="enc:5000.00",
                tipo="reajuste",
                motivo="Cadastro inicial",
                criado_em=datetime(2026, 1, 1, tzinfo=UTC),
            ),
            SimpleNamespace(
                data_vigencia=date(2026, 6, 1),
                salario_criptografado="enc:6000.00",
                tipo="reajuste",
                motivo="Promoção",
                criado_em=datetime(2026, 6, 1, tzinfo=UTC),
            ),
        ]

        with (
            patch("app.services.fichas_admissionais_service.users_service.buscar_usuario"),
            patch(
                "app.services.fichas_admissionais_service.historico_salarial_repository.listar_por_usuario",
                return_value=movimentos,
            ),
            patch(
                "app.services.fichas_admissionais_service.descriptografar_dado_sensivel",
                side_effect=lambda valor: valor.removeprefix("enc:") if valor else None,
            ),
        ):
            resultado = fichas_admissionais_service.historico_salarial(db, 7, SimpleNamespace(id=1))

        self.assertEqual(len(resultado), 2)
        self.assertEqual(resultado[0]["salario"], Decimal("5000.00"))
        self.assertEqual(resultado[1]["motivo"], "Promoção")
        self.assertEqual(db.added[0].acao, "HISTORICO_SALARIAL_CONSULTADO")
        self.assertEqual(db.commits, 1)

    def test_consultar_ficha_inexistente_retorna_none_e_audita(self):
        db = FakeDb()
        with (
            patch("app.services.fichas_admissionais_service.users_service.buscar_usuario"),
            patch(
                "app.services.fichas_admissionais_service.fichas_admissionais_repository.obter_por_usuario",
                return_value=None,
            ),
        ):
            resultado = fichas_admissionais_service.consultar_ficha(
                db,
                7,
                SimpleNamespace(id=1),
            )

        self.assertIsNone(resultado)
        self.assertEqual(db.added[0].acao, "FICHA_ADMISSIONAL_CONSULTADA")
        self.assertEqual(db.commits, 1)

    def test_modelo_xlsx_e_vinculado_ao_colaborador(self):
        db = FakeDb()
        user = SimpleNamespace(id=7, nome="Pessoa Teste", email="pessoa@empresa.com")
        with (
            patch("app.services.fichas_admissionais_service.users_service.buscar_usuario", return_value=user),
            patch(
                "app.services.fichas_admissionais_service.fichas_admissionais_repository.obter_por_usuario",
                return_value=None,
            ),
        ):
            conteudo = fichas_admissionais_service.gerar_modelo_xlsx(
                db,
                7,
                SimpleNamespace(id=1),
            )

        workbook = load_workbook(BytesIO(conteudo))
        planilha = workbook["Ficha admissional"]
        self.assertEqual(planilha["B2"].value, "Pessoa Teste")
        self.assertEqual(planilha["B3"].value, "pessoa@empresa.com")
        self.assertEqual(planilha.freeze_panes, "A6")
        campos = {planilha.cell(linha, 1).value for linha in range(1, planilha.max_row + 1)}
        self.assertIn("Número do RG", campos)
        self.assertIn("Salário", campos)
        self.assertEqual(db.added[0].acao, "MODELO_FICHA_ADMISSIONAL_EXPORTADO")

    def test_importar_xlsx_converte_campos_e_atualiza_ficha(self):
        db = FakeDb()
        user = SimpleNamespace(id=7, nome="Pessoa Teste", email="pessoa@empresa.com")
        with (
            patch("app.services.fichas_admissionais_service.users_service.buscar_usuario", return_value=user),
            patch(
                "app.services.fichas_admissionais_service.fichas_admissionais_repository.obter_por_usuario",
                return_value=None,
            ),
        ):
            modelo = fichas_admissionais_service.gerar_modelo_xlsx(
                db,
                7,
                SimpleNamespace(id=1),
            )

        workbook = load_workbook(BytesIO(modelo))
        planilha = workbook["Ficha admissional"]
        linhas = {
            planilha.cell(linha, 1).value: linha
            for linha in range(1, planilha.max_row + 1)
        }
        planilha.cell(linhas["Local do nascimento"], 2).value = "Brasília"
        planilha.cell(linhas["UF de nascimento"], 2).value = "DF"
        planilha.cell(linhas["Emissão do RG"], 2).value = date(2020, 5, 10)
        planilha.cell(linhas["Sexo"], 2).value = "Feminino"
        planilha.cell(linhas["Salário"], 2).value = "R$ 7.500,50"
        planilha.cell(linhas["Contrato de experiência (dias)"], 2).value = 90
        planilha.cell(linhas["Status da ficha"], 2).value = "Completa"
        arquivo = BytesIO()
        workbook.save(arquivo)

        ficha_resultado = {
            "id": 3,
            "user_id": 7,
            "status": "completa",
            "criado_em": datetime.now(UTC),
            "atualizado_em": datetime.now(UTC),
        }
        with (
            patch("app.services.fichas_admissionais_service.users_service.buscar_usuario", return_value=user),
            patch(
                "app.services.fichas_admissionais_service.atualizar_ficha",
                return_value=ficha_resultado,
            ) as atualizar,
        ):
            resultado = fichas_admissionais_service.importar_xlsx(
                db,
                7,
                "ficha.xlsx",
                arquivo.getvalue(),
                SimpleNamespace(id=1),
            )

        payload = atualizar.call_args.args[2]
        self.assertEqual(payload.local_nascimento, "Brasília")
        self.assertEqual(payload.rg_emissao, date(2020, 5, 10))
        self.assertEqual(payload.salario, Decimal("7500.50"))
        self.assertEqual(payload.contrato_experiencia_dias, 90)
        self.assertEqual(payload.status, "completa")
        self.assertEqual(resultado["ficha"], ficha_resultado)

    def test_importar_rejeita_planilha_de_outro_colaborador(self):
        db = FakeDb()
        user = SimpleNamespace(id=7, nome="Pessoa Teste", email="pessoa@empresa.com")
        with (
            patch("app.services.fichas_admissionais_service.users_service.buscar_usuario", return_value=user),
            patch(
                "app.services.fichas_admissionais_service.fichas_admissionais_repository.obter_por_usuario",
                return_value=None,
            ),
        ):
            modelo = fichas_admissionais_service.gerar_modelo_xlsx(
                db,
                7,
                SimpleNamespace(id=1),
            )
        workbook = load_workbook(BytesIO(modelo))
        workbook["Ficha admissional"]["B3"] = "outra@empresa.com"
        arquivo = BytesIO()
        workbook.save(arquivo)

        with (
            patch("app.services.fichas_admissionais_service.users_service.buscar_usuario", return_value=user),
            self.assertRaises(HTTPException) as exc,
        ):
            fichas_admissionais_service.importar_xlsx(
                db,
                7,
                "ficha.xlsx",
                arquivo.getvalue(),
                SimpleNamespace(id=1),
            )

        self.assertEqual(exc.exception.status_code, 400)
        self.assertIn("não corresponde", exc.exception.detail)


if __name__ == "__main__":
    unittest.main()
